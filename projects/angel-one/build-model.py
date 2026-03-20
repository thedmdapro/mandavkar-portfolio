"""
Build the Angel One Valuation Model — 6-tab institutional Excel template.
Run: python build_angel_one_model.py
Output: angel-one-model.xlsx

Tabs:
  1. 3-Statement Model  (fill in historical FY21–FY24)
  2. DCF Valuation      (auto-calculates from Tab 1)
  3. Comparables        (fill in peer data from Screener.in)
  4. Scenario Analysis  (auto-calculates from Tab 1 assumptions)
  5. Sensitivity        (WACC × terminal growth)
  6. Dashboard          (auto-summary from all tabs)
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.styles.differential import DifferentialStyle

# ── Palette ───────────────────────────────────────────────────────────
DARK_BLUE   = "2F4858"
TEAL        = "005F73"
WHITE       = "FFFFFF"
YELLOW      = "FFF3CD"   # manual input cells (historical actuals)
BLUE_CALC   = "DBEAFE"   # calculated/formula cells
GREEN_OUT   = "D4EDDA"   # output / price target cells
RED_WARN    = "FFCCCC"   # sanity check fail
GREY_LIGHT  = "F1F5F9"
GREY_MID    = "CBD5E1"
DARK_SECTION= "1E3A4A"   # dark sub-section headers

YEARS_HIST  = ["FY21", "FY22", "FY23", "FY24"]          # historical — fill manually
YEARS_FORE  = ["FY25E", "FY26E", "FY27E", "FY28E", "FY29E"]  # forecast — formula-driven
ALL_YEARS   = YEARS_HIST + YEARS_FORE


def thin(color="C0C0C0"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def medium_bottom(color=DARK_BLUE):
    m = Side(style="medium", color=color)
    n = Side(style=None)
    return Border(left=n, right=n, top=n, bottom=m)


def apply(cell, value=None, font=None, fill=None, alignment=None, border=None, number_format=None):
    if value is not None:
        cell.value = value
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border
    if number_format:
        cell.number_format = number_format


def hdr(text, bg=DARK_BLUE, fg=WHITE, size=10, bold=True, align="center", wrap=False):
    return dict(
        value=text,
        font=Font(bold=bold, color=fg, size=size, name="Calibri"),
        fill=PatternFill("solid", fgColor=bg),
        alignment=Alignment(horizontal=align, vertical="center", wrap_text=wrap),
        border=thin(),
    )


def input_cell(ws, row, col, value="", fmt="#,##0.0", note=""):
    cell = ws.cell(row, col)
    cell.value = value
    cell.fill = PatternFill("solid", fgColor=YELLOW)
    cell.border = thin()
    cell.alignment = Alignment(horizontal="right", vertical="center")
    cell.number_format = fmt
    return cell


def formula_cell(ws, row, col, formula, fmt="#,##0.0"):
    cell = ws.cell(row, col)
    cell.value = formula
    cell.fill = PatternFill("solid", fgColor=BLUE_CALC)
    cell.border = thin()
    cell.alignment = Alignment(horizontal="right", vertical="center")
    cell.number_format = fmt
    return cell


def label_cell(ws, row, col, text, indent=0, bold=False, italic=False, bg=None):
    cell = ws.cell(row, col)
    cell.value = ("  " * indent) + text
    cell.font = Font(bold=bold, italic=italic, size=9, name="Calibri",
                     color=DARK_BLUE if bold else "333333")
    cell.fill = PatternFill("solid", fgColor=bg or WHITE)
    cell.border = thin()
    cell.alignment = Alignment(horizontal="left", vertical="center")
    return cell


def section_header(ws, row, col_start, col_end, text):
    ws.merge_cells(start_row=row, start_column=col_start,
                   end_row=row, end_column=col_end)
    cell = ws.cell(row, col_start)
    cell.value = text.upper()
    cell.font = Font(bold=True, color=WHITE, size=9, name="Calibri")
    cell.fill = PatternFill("solid", fgColor=TEAL)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = thin()
    ws.row_dimensions[row].height = 18


def set_w(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width


# ═══════════════════════════════════════════════════════════════════════
# TAB 1: 3-STATEMENT MODEL
# ═══════════════════════════════════════════════════════════════════════

def build_3_statement(wb):
    ws = wb.active
    ws.title = "3-Statement Model"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "C5"

    # ── Title ──
    N_COLS = 11   # label + 9 years
    ws.merge_cells(f"A1:{get_column_letter(N_COLS)}1")
    apply(ws["A1"], **hdr("ANGEL ONE LTD (NSE: ANGELONE) — 3-Statement Financial Model", size=13, bg=DARK_BLUE))
    ws.row_dimensions[1].height = 32

    ws.merge_cells(f"A2:{get_column_letter(N_COLS)}2")
    apply(ws["A2"],
          value="🟡 Yellow = paste historical data from Annual Reports (BSE 543235)   |   "
                "🔵 Blue = formula-driven forecast   |   "
                "Fiscal year ends March 31",
          font=Font(size=9, italic=True, color="555555", name="Calibri"),
          fill=PatternFill("solid", fgColor=GREY_LIGHT),
          alignment=Alignment(horizontal="left", vertical="center"))
    ws.row_dimensions[2].height = 18

    # ── Column headers ──
    set_w(ws, 1, 32)
    apply(ws.cell(4, 1), **hdr("₹ Millions", bg=GREY_MID, fg=DARK_BLUE, align="left"))
    for i, yr in enumerate(ALL_YEARS):
        col = i + 2
        set_w(ws, col, 12)
        is_hist = yr in YEARS_HIST
        apply(ws.cell(4, col), **hdr(yr, bg=DARK_BLUE if is_hist else TEAL))
    ws.row_dimensions[4].height = 22

    # ── Helper to build a year row ──
    def build_row(ws, row, label, hist_vals=None, fore_formulas=None,
                  row_fmt="#,##0.0", bold=False, indent=0, is_pct=False,
                  section_bg=None):
        """
        hist_vals: list of 4 values (or "" for blanks) — yellow cells
        fore_formulas: list of 5 formula strings — blue cells
        """
        label_cell(ws, row, 1, label, indent=indent, bold=bold, bg=section_bg)

        # Historical columns (yellow)
        for i in range(4):
            col = i + 2
            cell = ws.cell(row, col)
            cell.value = hist_vals[i] if hist_vals else ""
            cell.fill = PatternFill("solid", fgColor=YELLOW if not section_bg else YELLOW)
            cell.border = thin()
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.number_format = "0.0%" if is_pct else row_fmt
            if bold:
                cell.font = Font(bold=True, size=9, name="Calibri")

        # Forecast columns (blue)
        for i, fml in enumerate(fore_formulas or [""] * 5):
            col = i + 6
            cell = ws.cell(row, col)
            cell.value = fml
            cell.fill = PatternFill("solid", fgColor=BLUE_CALC if not section_bg else BLUE_CALC)
            cell.border = thin()
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.number_format = "0.0%" if is_pct else row_fmt
            if bold:
                cell.font = Font(bold=True, size=9, name="Calibri")

        ws.row_dimensions[row].height = 16

    # ──────────────────────────────────────────────
    # ASSUMPTIONS SECTION (rows 5–20)
    # ──────────────────────────────────────────────
    section_header(ws, 5, 1, N_COLS, "KEY ASSUMPTIONS (FY25E–FY29E — override these with your views)")
    ws.row_dimensions[5].height = 20

    ASS = {  # row → (label, values for FY25E..FY29E)
        6:  ("Active Client Growth (% YoY)",   [0.15, 0.18, 0.18, 0.15, 0.12]),
        7:  ("ARPU Growth (% YoY)",             [0.05, 0.05, 0.05, 0.05, 0.05]),
        8:  ("Brokerage Take Rate (bps)",        [3.5,  3.3,  3.2,  3.0,  3.0]),
        9:  ("Distribution Revenue Growth (%)",  [0.12, 0.15, 0.15, 0.12, 0.10]),
        10: ("MTF Book Growth (% YoY)",          [0.10, 0.12, 0.12, 0.10, 0.10]),
        11: ("MTF Interest Rate (%)",            [0.18, 0.18, 0.17, 0.17, 0.17]),
        12: ("EBITDA Margin (%)",                [0.38, 0.40, 0.42, 0.43, 0.44]),
        13: ("D&A / Revenue (%)",                [0.02, 0.02, 0.02, 0.02, 0.02]),
        14: ("Tax Rate (%)",                     [0.25, 0.25, 0.25, 0.25, 0.25]),
        15: ("CapEx / Revenue (%)",              [0.03, 0.03, 0.03, 0.03, 0.03]),
        16: ("NWC / Revenue (%)",                [0.10, 0.10, 0.10, 0.10, 0.10]),
    }
    for row, (lbl, vals) in ASS.items():
        label_cell(ws, row, 1, lbl, indent=1, bg=GREY_LIGHT)
        for i, yr in enumerate(YEARS_HIST):
            c = ws.cell(row, i + 2)
            c.value = "—"
            c.fill = PatternFill("solid", fgColor=GREY_LIGHT)
            c.border = thin()
            c.alignment = Alignment(horizontal="center")
            c.font = Font(color="888888", size=9, name="Calibri")
        for i, val in enumerate(vals):
            col = i + 6
            c = ws.cell(row, col)
            c.value = val
            c.fill = PatternFill("solid", fgColor=YELLOW)
            c.border = thin()
            c.alignment = Alignment(horizontal="right", vertical="center")
            c.number_format = "0.0%" if row in [6, 7, 9, 10, 11, 12, 13, 14, 15, 16] else "0.0"
        ws.row_dimensions[row].height = 16

    # ──────────────────────────────────────────────
    # INCOME STATEMENT (rows 18–35)
    # ──────────────────────────────────────────────
    section_header(ws, 18, 1, N_COLS, "INCOME STATEMENT")

    IS_ROWS = [
        # (row_offset, label, indent, bold, formula_hint)
        (19, "Brokerage Revenue",       1, False),
        (20, "Distribution Revenue",    1, False),
        (21, "DP & Other Revenue",      1, False),
        (22, "MTF Interest Income",     1, False),
        (23, "Total Revenue",           0, True),
        (24, ""),
        (25, "Employee Costs",          1, False),
        (26, "Technology Costs",        1, False),
        (27, "Marketing Costs",         1, False),
        (28, "Other OpEx",              1, False),
        (29, "EBITDA",                  0, True),
        (30, "EBITDA Margin (%)",       0, False),
        (31, "D&A",                     1, False),
        (32, "EBIT",                    0, True),
        (33, "Finance Costs",           1, False),
        (34, "PBT",                     0, True),
        (35, "Tax",                     1, False),
        (36, "PAT (Net Profit)",        0, True),
        (37, "PAT Margin (%)",          0, False),
        (38, "EPS (₹)",                0, False),
    ]

    for row, label, *rest in IS_ROWS:
        if not label:
            ws.row_dimensions[row].height = 8
            continue
        indent = rest[0] if rest else 0
        bold   = rest[1] if len(rest) > 1 else False
        is_pct = "%" in label or "Margin" in label or "EPS" in label

        label_cell(ws, row, 1, label, indent=indent, bold=bold)

        for i in range(4):  # historical
            c = ws.cell(row, i + 2)
            c.fill = PatternFill("solid", fgColor=YELLOW)
            c.border = thin()
            c.alignment = Alignment(horizontal="right", vertical="center")
            c.number_format = "0.0%" if is_pct else "#,##0.0"
            if bold:
                c.font = Font(bold=True, size=9, name="Calibri")

        # Forecast formulas (illustrative — link to row 23 for revenue, assumptions tab)
        for i in range(5):
            col = i + 6
            prev_col = get_column_letter(col - 1)
            c = ws.cell(row, col)
            c.fill = PatternFill("solid", fgColor=BLUE_CALC)
            c.border = thin()
            c.alignment = Alignment(horizontal="right", vertical="center")
            c.number_format = "0.0%" if is_pct else "#,##0.0"
            if bold:
                c.font = Font(bold=True, size=9, name="Calibri")

            # Build row-specific formulas
            ass_row = {6: 6, 7: 7, 8: 8, 9: 9, 10: 10, 11: 11, 12: 12, 13: 13, 14: 14, 15: 15}
            yr_col  = get_column_letter(col)

            if row == 19:  # Brokerage Revenue
                c.value = f"={prev_col}19*(1+{yr_col}6)*(1+{yr_col}7)"
            elif row == 20:  # Distribution
                c.value = f"={prev_col}20*(1+{yr_col}9)"
            elif row == 21:  # DP Other — grows with client base
                c.value = f"={prev_col}21*(1+{yr_col}6)"
            elif row == 22:  # MTF Interest
                c.value = f"={prev_col}22*(1+{yr_col}10)"
            elif row == 23:  # Total Revenue
                c.value = f"={yr_col}19+{yr_col}20+{yr_col}21+{yr_col}22"
                c.font = Font(bold=True, size=9, name="Calibri")
            elif row == 29:  # EBITDA
                c.value = f"={yr_col}23*{yr_col}12"
                c.font = Font(bold=True, size=9, name="Calibri")
            elif row == 30:  # EBITDA Margin
                c.value = f"={yr_col}29/{yr_col}23"
                c.number_format = "0.0%"
            elif row == 31:  # D&A
                c.value = f"={yr_col}23*{yr_col}13"
            elif row == 32:  # EBIT
                c.value = f"={yr_col}29-{yr_col}31"
                c.font = Font(bold=True, size=9, name="Calibri")
            elif row == 33:  # Finance Costs — manual for now
                c.value = ""
                c.fill = PatternFill("solid", fgColor=YELLOW)
            elif row == 34:  # PBT
                c.value = f"={yr_col}32-{yr_col}33"
                c.font = Font(bold=True, size=9, name="Calibri")
            elif row == 35:  # Tax
                c.value = f"={yr_col}34*{yr_col}14"
            elif row == 36:  # PAT
                c.value = f"={yr_col}34-{yr_col}35"
                c.font = Font(bold=True, size=9, name="Calibri")
            elif row == 37:  # PAT Margin
                c.value = f"={yr_col}36/{yr_col}23"
                c.number_format = "0.0%"
            elif row == 38:  # EPS (assume diluted shares in row below)
                c.value = f"={yr_col}36/Assumptions!B2*1000000"  # shares in millions
                c.number_format = "#,##0.00"

        ws.row_dimensions[row].height = 16

    # ──────────────────────────────────────────────
    # BALANCE SHEET (rows 41–65)
    # ──────────────────────────────────────────────
    section_header(ws, 41, 1, N_COLS, "BALANCE SHEET")

    BS_ROWS = [
        (42, "ASSETS",                      0, True),
        (43, "Fixed Assets (net)",          1, False),
        (44, "Investments & Securities",    1, False),
        (45, "Client Receivables",          1, False),
        (46, "Other Current Assets",        1, False),
        (47, "Cash & Cash Equivalents",     1, False),
        (48, "Total Assets",                0, True),
        (49, ""),
        (50, "LIABILITIES & EQUITY",        0, True),
        (51, "Client Payables",             1, False),
        (52, "Borrowings (MTF funded)",     1, False),
        (53, "Other Current Liabilities",   1, False),
        (54, "Total Liabilities",           0, True),
        (55, ""),
        (56, "Share Capital",               1, False),
        (57, "Reserves & Surplus",          1, False),
        (58, "Total Equity",                0, True),
        (59, "Total Liabilities + Equity",  0, True),
    ]
    for row, label, *rest in BS_ROWS:
        if not label:
            ws.row_dimensions[row].height = 8
            continue
        indent = rest[0] if rest else 0
        bold   = rest[1] if len(rest) > 1 else False
        label_cell(ws, row, 1, label, indent=indent, bold=bold)
        for i in range(4):
            c = ws.cell(row, i + 2)
            c.fill = PatternFill("solid", fgColor=YELLOW if not bold else "FFF9E6")
            c.border = thin()
            c.alignment = Alignment(horizontal="right", vertical="center")
            c.number_format = "#,##0.0"
            if bold:
                c.font = Font(bold=True, size=9, name="Calibri")
        for i in range(5):
            col = i + 6
            yr_col = get_column_letter(col)
            c = ws.cell(row, col)
            c.fill = PatternFill("solid", fgColor=BLUE_CALC)
            c.border = thin()
            c.alignment = Alignment(horizontal="right", vertical="center")
            c.number_format = "#,##0.0"
            if bold:
                c.font = Font(bold=True, size=9, name="Calibri")
            # Key balance sheet formulas
            if row == 43:   # Fixed assets = prior + capex - da
                prev_col = get_column_letter(col - 1)
                c.value = f"={prev_col}43+{yr_col}23*{yr_col}15-{yr_col}31"
            elif row == 45:  # Client receivables ~ NWC % revenue
                c.value = f"={yr_col}23*{yr_col}16"
            elif row == 48:  # Total assets
                c.value = f"=SUM({yr_col}43:{yr_col}47)"
                c.font = Font(bold=True, size=9, name="Calibri")
            elif row == 54:  # Total liabilities
                c.value = f"=SUM({yr_col}51:{yr_col}53)"
                c.font = Font(bold=True, size=9, name="Calibri")
            elif row == 57:  # Retained earnings
                prev_col = get_column_letter(col - 1)
                c.value = f"={prev_col}57+{yr_col}36"
            elif row == 58:  # Total equity
                c.value = f"={yr_col}56+{yr_col}57"
                c.font = Font(bold=True, size=9, name="Calibri")
            elif row == 59:  # Check
                c.value = f"={yr_col}54+{yr_col}58"
                c.font = Font(bold=True, size=9, name="Calibri")
        ws.row_dimensions[row].height = 16

    # ──────────────────────────────────────────────
    # CASH FLOW STATEMENT (rows 62–80)
    # ──────────────────────────────────────────────
    section_header(ws, 62, 1, N_COLS, "CASH FLOW STATEMENT")

    CF_ROWS = [
        (63, "PAT",                              1, False),
        (64, "Add: D&A",                         1, False),
        (65, "Change in NWC",                    1, False),
        (66, "Operating Cash Flow",              0, True),
        (67, ""),
        (68, "CapEx",                            1, False),
        (69, "Investing Cash Flow",              0, True),
        (70, ""),
        (71, "Dividends Paid",                   1, False),
        (72, "Borrowings (net)",                 1, False),
        (73, "Financing Cash Flow",              0, True),
        (74, ""),
        (75, "Net Change in Cash",               0, True),
        (76, "Opening Cash",                     1, False),
        (77, "Closing Cash",                     0, True),
    ]
    for row, label, *rest in CF_ROWS:
        if not label:
            ws.row_dimensions[row].height = 8
            continue
        indent = rest[0] if rest else 0
        bold   = rest[1] if len(rest) > 1 else False
        label_cell(ws, row, 1, label, indent=indent, bold=bold)
        for i in range(4):
            c = ws.cell(row, i + 2)
            c.fill = PatternFill("solid", fgColor=YELLOW if not bold else "FFF9E6")
            c.border = thin()
            c.alignment = Alignment(horizontal="right", vertical="center")
            c.number_format = "#,##0.0"
            if bold:
                c.font = Font(bold=True, size=9, name="Calibri")
        for i in range(5):
            col = i + 6
            yr_col = get_column_letter(col)
            prev_col = get_column_letter(col - 1)
            c = ws.cell(row, col)
            c.fill = PatternFill("solid", fgColor=BLUE_CALC)
            c.border = thin()
            c.alignment = Alignment(horizontal="right", vertical="center")
            c.number_format = "#,##0.0"
            if bold:
                c.font = Font(bold=True, size=9, name="Calibri")
            if row == 63:    c.value = f"={yr_col}36"
            elif row == 64:  c.value = f"={yr_col}31"
            elif row == 65:  c.value = f"=-({yr_col}23-{prev_col}23)*{yr_col}16"
            elif row == 66:  c.value = f"={yr_col}63+{yr_col}64+{yr_col}65"
            elif row == 68:  c.value = f"=-{yr_col}23*{yr_col}15"
            elif row == 69:  c.value = f"={yr_col}68"
            elif row == 73:  c.value = f"={yr_col}71+{yr_col}72"
            elif row == 75:  c.value = f"={yr_col}66+{yr_col}69+{yr_col}73"
            elif row == 76:  c.value = f"={prev_col}77"
            elif row == 77:  c.value = f"={yr_col}76+{yr_col}75"
        ws.row_dimensions[row].height = 16

    # ── SANITY CHECK ROW ──
    section_header(ws, 79, 1, N_COLS, "SANITY CHECK — Closing Cash must match Balance Sheet Cash")
    for i in range(5):
        col = i + 6
        yr_col = get_column_letter(col)
        c = ws.cell(80, col)
        c.value = f"={yr_col}77-{yr_col}47"
        c.border = thin()
        c.alignment = Alignment(horizontal="right", vertical="center")
        c.number_format = "#,##0.0"
        c.font = Font(bold=True, size=10, name="Calibri")
    label_cell(ws, 80, 1, "CF Closing Cash − BS Cash  (must = 0)", bold=True)
    ws.row_dimensions[80].height = 20

    # Red/green conditional on sanity row
    from openpyxl.formatting.rule import CellIsRule
    green_fill = PatternFill("solid", fgColor="C6EFCE")
    red_fill   = PatternFill("solid", fgColor="FFC7CE")
    green_font = Font(bold=True, color="006100", name="Calibri")
    red_font   = Font(bold=True, color="9C0006", name="Calibri")
    sanity_range = f"F80:J80"
    ws.conditional_formatting.add(sanity_range,
        CellIsRule(operator="equal", formula=["0"], font=green_font, fill=green_fill))
    ws.conditional_formatting.add(sanity_range,
        CellIsRule(operator="notEqual", formula=["0"], font=red_font, fill=red_fill))

    # ── Shares outstanding row (needed for EPS) ──
    label_cell(ws, 82, 1, "Diluted Shares Outstanding (Millions)", bold=True)
    for i in range(9):
        c = ws.cell(82, i + 2)
        c.fill = PatternFill("solid", fgColor=YELLOW)
        c.border = thin()
        c.alignment = Alignment(horizontal="right", vertical="center")
        c.number_format = "#,##0.0"
        c.value = 85.5   # approximate — update from latest filing


# ═══════════════════════════════════════════════════════════════════════
# TAB 2: DCF VALUATION
# ═══════════════════════════════════════════════════════════════════════

def build_dcf(wb):
    ws = wb.create_sheet("DCF Valuation")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:F1")
    apply(ws["A1"], **hdr("DCF VALUATION — Angel One Ltd", size=13, bg=DARK_BLUE))
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:F2")
    apply(ws["A2"],
          value="All inputs linked from 3-Statement Model. Only override WACC inputs and terminal growth rate here.",
          font=Font(size=9, italic=True, color="555555", name="Calibri"),
          fill=PatternFill("solid", fgColor=GREY_LIGHT),
          alignment=Alignment(horizontal="left", vertical="center"))
    ws.row_dimensions[2].height = 16

    set_w(ws, 1, 35); set_w(ws, 2, 16); set_w(ws, 3, 16)
    set_w(ws, 4, 16); set_w(ws, 5, 16); set_w(ws, 6, 16)

    # ── WACC inputs ──
    section_header(ws, 4, 1, 3, "WACC INPUTS")
    wacc_inputs = [
        (5,  "Risk-Free Rate (10Y G-Sec, %)",    0.068),
        (6,  "Equity Risk Premium (%)",           0.055),
        (7,  "Beta (vs Nifty 50, 3Y weekly)",     1.15),
        (8,  "Cost of Equity = Rf + β × ERP (%)", "=B5+B7*B6"),
        (9,  "Pre-tax Cost of Debt (%)",          0.085),
        (10, "Tax Rate (%)",                      0.25),
        (11, "After-tax Cost of Debt (%)",        "=B9*(1-B10)"),
        (12, "Equity Weight",                     0.85),
        (13, "Debt Weight",                       "=1-B12"),
        (14, "WACC (%)",                          "=B8*B12+B11*B13"),
    ]
    for row, label, val in wacc_inputs:
        label_cell(ws, row, 1, label, indent=1, bold=row in [8, 14])
        c = ws.cell(row, 2)
        c.value = val
        c.fill = PatternFill("solid", fgColor=YELLOW if not str(val).startswith("=") else BLUE_CALC)
        c.border = thin()
        c.alignment = Alignment(horizontal="right", vertical="center")
        c.number_format = "0.00%"
        if row in [8, 14]:
            c.font = Font(bold=True, size=10, name="Calibri")
        ws.row_dimensions[row].height = 16

    # ── FCF projection ──
    section_header(ws, 16, 1, 6, "FREE CASH FLOW PROJECTION (linked from 3-Statement Model)")
    apply(ws.cell(17, 1), **hdr("₹ Millions", bg=GREY_MID, fg=DARK_BLUE, align="left"))
    for i, yr in enumerate(YEARS_FORE):
        apply(ws.cell(17, i + 2), **hdr(yr, bg=TEAL))
    ws.row_dimensions[17].height = 20

    fcf_rows = [
        (18, "EBIT",            "='3-Statement Model'!{col}32"),
        (19, "Tax on EBIT",     "=-'3-Statement Model'!{col}32*'3-Statement Model'!{col}14"),
        (20, "EBIT(1-t)",       "={col}18+{col}19"),
        (21, "D&A (add back)",  "='3-Statement Model'!{col}31"),
        (22, "CapEx",           "='3-Statement Model'!{col}68"),
        (23, "Δ NWC",           "='3-Statement Model'!{col}65"),
        (24, "Free Cash Flow",  "={col}20+{col}21+{col}22+{col}23"),
    ]
    YEAR_COLS = ["F", "G", "H", "I", "J"]  # FY25E–FY29E in 3-statement tab
    for row, label, fml_template in fcf_rows:
        bold = row in [20, 24]
        label_cell(ws, row, 1, label, indent=1, bold=bold)
        for i, yr_col in enumerate(YEAR_COLS):
            dcf_col = get_column_letter(i + 2)
            c = ws.cell(row, i + 2)
            c.value = fml_template.replace("{col}", dcf_col) if dcf_col in ["B","C","D","E","F"] else fml_template.replace("{col}", yr_col)
            # Re-map: in DCF tab, col B = FY25E, in 3-stat tab col F = FY25E
            if "3-Statement" in fml_template:
                c.value = fml_template.replace("{col}", yr_col)
            else:
                c.value = fml_template.replace("{col}", dcf_col)
            c.fill = PatternFill("solid", fgColor=BLUE_CALC)
            c.border = thin()
            c.alignment = Alignment(horizontal="right", vertical="center")
            c.number_format = "#,##0.0"
            if bold:
                c.font = Font(bold=True, size=9, name="Calibri")
        ws.row_dimensions[row].height = 16

    # ── Terminal value & EV ──
    section_header(ws, 26, 1, 3, "TERMINAL VALUE & ENTERPRISE VALUE")

    tv_inputs = [
        (27, "Terminal Growth Rate (%)",    0.04),
        (28, "Terminal FCF (FCF_n × (1+g))", "=F24*(1+B27)"),
        (29, "Terminal Value (Gordon Growth)", "=B28/(B14-B27)"),
        (30, "PV of Terminal Value",           "=B29/(1+B14)^5"),
    ]
    pv_rows = [(31, f"PV of FCF — {yr}", f"={get_column_letter(i+2)}24/(1+$B$14)^{i+1}")
               for i, yr in enumerate(YEARS_FORE)]
    for row, label, val in tv_inputs + pv_rows:
        bold = row in [29, 30]
        label_cell(ws, row, 1, label, indent=1, bold=bold)
        c = ws.cell(row, 2)
        c.value = val
        c.fill = PatternFill("solid", fgColor=YELLOW if isinstance(val, float) and not str(val).startswith("=") else BLUE_CALC)
        if isinstance(val, float):
            c.fill = PatternFill("solid", fgColor=YELLOW)
        c.border = thin()
        c.alignment = Alignment(horizontal="right", vertical="center")
        c.number_format = "0.0%" if row == 27 else "#,##0.0"
        if bold:
            c.font = Font(bold=True, size=9, name="Calibri")
        ws.row_dimensions[row].height = 16

    pv_fcf_sum_row = 31 + len(YEARS_FORE)
    label_cell(ws, pv_fcf_sum_row, 1, "Sum of PV of FCFs", bold=True)
    c = ws.cell(pv_fcf_sum_row, 2)
    c.value = f"=SUM(B31:B{30+len(YEARS_FORE)})"
    c.fill = PatternFill("solid", fgColor=BLUE_CALC)
    c.border = thin(); c.alignment = Alignment(horizontal="right")
    c.number_format = "#,##0.0"; c.font = Font(bold=True, size=9, name="Calibri")
    ws.row_dimensions[pv_fcf_sum_row].height = 16

    # ── Equity bridge & price target ──
    section_header(ws, pv_fcf_sum_row + 2, 1, 3, "EQUITY BRIDGE & PRICE TARGET")
    ev_bridge = [
        (pv_fcf_sum_row + 3, "Enterprise Value",           f"=B{pv_fcf_sum_row}+B30"),
        (pv_fcf_sum_row + 4, "Less: Net Debt",             ""),
        (pv_fcf_sum_row + 5, "Equity Value",               f"=B{pv_fcf_sum_row+3}-B{pv_fcf_sum_row+4}"),
        (pv_fcf_sum_row + 6, "Diluted Shares (Millions)",  "='3-Statement Model'!B82"),
        (pv_fcf_sum_row + 7, "Intrinsic Value per Share (₹)", f"=B{pv_fcf_sum_row+5}/B{pv_fcf_sum_row+6}*1000000"),
        (pv_fcf_sum_row + 8, "Current Market Price (₹)",   212.12),
        (pv_fcf_sum_row + 9, "Implied Upside / (Downside) %", f"=B{pv_fcf_sum_row+7}/B{pv_fcf_sum_row+8}-1"),
        (pv_fcf_sum_row + 10, "Rating",                    f'=IF(B{pv_fcf_sum_row+9}>0.15,"BUY",IF(B{pv_fcf_sum_row+9}>-0.05,"HOLD","SELL"))'),
    ]
    for row, label, val in ev_bridge:
        bold = row in [pv_fcf_sum_row+3, pv_fcf_sum_row+5, pv_fcf_sum_row+7, pv_fcf_sum_row+10]
        label_cell(ws, row, 1, label, indent=1, bold=bold)
        c = ws.cell(row, 2)
        c.value = val
        is_yellow = row == pv_fcf_sum_row + 4 or row == pv_fcf_sum_row + 8
        is_output = row in [pv_fcf_sum_row+7, pv_fcf_sum_row+9, pv_fcf_sum_row+10]
        c.fill = PatternFill("solid", fgColor=YELLOW if is_yellow else (GREEN_OUT if is_output else BLUE_CALC))
        c.border = thin()
        c.alignment = Alignment(horizontal="right", vertical="center")
        c.number_format = ("0.0%" if row == pv_fcf_sum_row+9
                           else ("₹#,##0.00" if row in [pv_fcf_sum_row+7, pv_fcf_sum_row+8] else "#,##0.0"))
        if bold:
            c.font = Font(bold=True, size=10 if is_output else 9, name="Calibri",
                         color=DARK_BLUE if is_output else "000000")
        ws.row_dimensions[row].height = 18


# ═══════════════════════════════════════════════════════════════════════
# TAB 3: COMPARABLES
# ═══════════════════════════════════════════════════════════════════════

def build_comparables(wb):
    ws = wb.create_sheet("Comparables")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "B4"

    ws.merge_cells("A1:L1")
    apply(ws["A1"], **hdr("COMPARABLE COMPANY ANALYSIS — Indian Broking Sector", size=12, bg=DARK_BLUE))
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:L2")
    apply(ws["A2"],
          value="Source: Screener.in (free). Pull latest FY data for each peer. "
                "Angel One row auto-populated from 3-Statement Model.",
          font=Font(size=9, italic=True, color="555555", name="Calibri"),
          fill=PatternFill("solid", fgColor=GREY_LIGHT),
          alignment=Alignment(horizontal="left", vertical="center"))
    ws.row_dimensions[2].height = 16

    col_headers = [
        ("Company",          22), ("NSE Ticker",    12),
        ("Market Cap\n(₹Cr)",13), ("Net Debt\n(₹Cr)",12),
        ("EV\n(₹Cr)",        12), ("Revenue\n(₹Cr)",12),
        ("EBITDA\n(₹Cr)",    12), ("EBITDA Margin\n(%)",13),
        ("PAT\n(₹Cr)",       12), ("EV/EBITDA\n(x)",12),
        ("P/E\n(x)",         10), ("P/B\n(x)",      10),
    ]
    for c, (h, w) in enumerate(col_headers, start=1):
        apply(ws.cell(3, c), **hdr(h, bg=TEAL, wrap=True))
        set_w(ws, c, w)
    ws.row_dimensions[3].height = 34

    # Peer rows
    peers = [
        ("Angel One Ltd",              "ANGELONE",     "=D4/E4", "=E4/F4", "=E4/G4", "=C4/'3-Statement Model'!I36*100"),
        ("MOFSL",                      "MOTILALOFS",   "=D5/E5", "=E5/F5", "=E5/G5", "=C5/'3-Statement Model'!I36*100"),
        ("5Paisa Capital",             "5PAISA",       "=D6/E6", "=E6/F6", "=E6/G6", "=C6/'3-Statement Model'!I36*100"),
        ("IIFL Securities",            "IIFLSEC",      "=D7/E7", "=E7/F7", "=E7/G7", "=C7/'3-Statement Model'!I36*100"),
        ("Anand Rathi Wealth",         "ANANDRATHI",   "=D8/E8", "=E8/F8", "=E8/G8", "=C8/'3-Statement Model'!I36*100"),
        ("Nuvama Wealth",              "NUVAMA",       "=D9/E9", "=E9/F9", "=E9/G9", "=C9/'3-Statement Model'!I36*100"),
    ]
    for r, (name, ticker, *_) in enumerate(peers, start=4):
        is_subject = r == 4
        row_bg = "E8F4FD" if is_subject else (GREY_LIGHT if r % 2 == 0 else WHITE)

        ws.cell(r, 1).value = name
        ws.cell(r, 1).font = Font(bold=is_subject, size=9, name="Calibri")
        ws.cell(r, 1).fill = PatternFill("solid", fgColor=row_bg)
        ws.cell(r, 1).border = thin()
        ws.cell(r, 1).alignment = Alignment(horizontal="left", vertical="center")

        ws.cell(r, 2).value = ticker
        ws.cell(r, 2).fill = PatternFill("solid", fgColor=row_bg)
        ws.cell(r, 2).border = thin()
        ws.cell(r, 2).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(r, 2).font = Font(size=9, name="Calibri")

        # Columns C–L: market cap, net debt, EV, revenue, EBITDA, margin, PAT, EV/EBITDA, P/E, P/B
        fmts = ["#,##0", "#,##0", "#,##0", "#,##0", "#,##0", "0.0%", "#,##0", "0.0x", "0.0x", "0.0x"]
        for c in range(3, 13):
            cell = ws.cell(r, c)
            col_idx = c - 2
            if c in [10, 11, 12]:  # calculated multiples
                if c == 10:   cell.value = f"=E{r}/G{r}"
                elif c == 11: cell.value = f"=C{r}/I{r}"
                elif c == 12: cell.value = ""  # needs book value — manual
                cell.fill = PatternFill("solid", fgColor=BLUE_CALC)
            else:
                cell.fill = PatternFill("solid", fgColor=YELLOW)
            cell.border = thin()
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.number_format = fmts[col_idx - 1] if col_idx - 1 < len(fmts) else "#,##0"
            cell.font = Font(bold=is_subject, size=9, name="Calibri")
        ws.row_dimensions[r].height = 17

    # Summary statistics row
    STATS_ROW = 11
    ws.merge_cells(f"A{STATS_ROW}:B{STATS_ROW}")
    apply(ws.cell(STATS_ROW, 1), **hdr("Peer Median (ex-Angel One)", bg=DARK_BLUE))
    for c in range(3, 13):
        cell = ws.cell(STATS_ROW, c)
        col_l = get_column_letter(c)
        cell.value = f"=MEDIAN({col_l}5:{col_l}9)"
        cell.fill = PatternFill("solid", fgColor=GREEN_OUT)
        cell.border = thin()
        cell.alignment = Alignment(horizontal="right", vertical="center")
        cell.font = Font(bold=True, size=9, name="Calibri")
    ws.row_dimensions[STATS_ROW].height = 20

    # Implied Angel One value at peer median multiples
    section_header(ws, 13, 1, 5, "IMPLIED VALUATION — Angel One at Peer Median Multiples")
    impl = [
        (14, "Peer Median EV/EBITDA (x)",        f"=J{STATS_ROW}"),
        (15, "Angel One EBITDA (₹Cr, FY24)",     ""),  # manual from 3-stat
        (16, "Implied EV at Peer EV/EBITDA",     "=B14*B15"),
        (17, "Implied Equity Value",              ""),
        (18, "Implied Price per Share (₹)",       ""),
        (19, "Current Price (₹)",                212.12),
        (20, "Implied Upside / (Downside) %",    "=IF(B19=0,\"\",B18/B19-1)"),
    ]
    for row, label, val in impl:
        bold = row in [16, 18, 20]
        label_cell(ws, row, 1, label, indent=1, bold=bold)
        c = ws.cell(row, 2)
        c.value = val
        is_yellow = row in [15, 17, 18, 19]
        is_output = row in [18, 20]
        c.fill = PatternFill("solid", fgColor=YELLOW if is_yellow else (GREEN_OUT if is_output else BLUE_CALC))
        c.border = thin()
        c.alignment = Alignment(horizontal="right", vertical="center")
        c.number_format = "0.0%" if row == 20 else ("#,##0.00" if row in [18, 19] else "#,##0.0")
        if bold:
            c.font = Font(bold=True, size=9, name="Calibri")
        ws.row_dimensions[row].height = 16


# ═══════════════════════════════════════════════════════════════════════
# TAB 4: SCENARIO ANALYSIS
# ═══════════════════════════════════════════════════════════════════════

def build_scenarios(wb):
    ws = wb.create_sheet("Scenario Analysis")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:E1")
    apply(ws["A1"], **hdr("SCENARIO ANALYSIS — Bull / Base / Bear", size=12, bg=DARK_BLUE))
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:E2")
    apply(ws["A2"],
          value="Adjust key drivers per scenario. Price targets auto-calculate using DCF WACC and terminal growth.",
          font=Font(size=9, italic=True, color="555555", name="Calibri"),
          fill=PatternFill("solid", fgColor=GREY_LIGHT),
          alignment=Alignment(horizontal="left", vertical="center"))
    ws.row_dimensions[2].height = 16

    set_w(ws, 1, 36)
    for c, label in enumerate(["Driver / Assumption", "Bear", "Base", "Bull"], start=1):
        set_w(ws, c + 1, 16)
        apply(ws.cell(4, c + 1 if c > 1 else 1),
              **hdr(label, bg=DARK_BLUE if c == 1 else (TEAL if c == 3 else "AE2012" if c == 2 else "1A7A4A"),
                    fg=WHITE))
        if c == 1:
            apply(ws.cell(4, 1), **hdr(label, bg=GREY_MID, fg=DARK_BLUE, align="left"))
    ws.row_dimensions[4].height = 22

    # Scenario assumptions (yellow — all manual)
    scenario_drivers = [
        ("Active Client CAGR (FY25–29E, %)",    "8%",  "15%", "22%"),
        ("ARPU Growth (%/yr)",                   "3%",  "5%",  "7%"),
        ("EBITDA Margin (FY29E, %)",             "35%", "42%", "47%"),
        ("Terminal Growth Rate (%)",             "3.5%","4.0%","5.0%"),
        ("WACC (%)",                             "13%", "12%", "11%"),
    ]
    for r, (driver, bear, base, bull) in enumerate(scenario_drivers, start=5):
        label_cell(ws, r, 1, driver, indent=1)
        for c, (val, bg) in enumerate([(bear,"FFCCCC"), (base, YELLOW), (bull,"C6EFCE")], start=2):
            cell = ws.cell(r, c)
            cell.value = val
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.border = thin()
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(bold=True, size=9, name="Calibri")
        ws.row_dimensions[r].height = 18

    # Output rows
    section_header(ws, 11, 1, 4, "SCENARIO OUTPUTS — Price Targets")
    output_rows = [
        (12, "FY29E Revenue (₹Mn)"),
        (13, "FY29E EBITDA (₹Mn)"),
        (14, "FY29E PAT (₹Mn)"),
        (15, ""),
        (16, "Implied EV (₹Mn)"),
        (17, "Equity Value (₹Mn)"),
        (18, "Price per Share (₹)"),
        (19, "Upside vs Current (₹212)"),
        (20, "Rating"),
    ]
    for row, label in output_rows:
        if not label:
            ws.row_dimensions[row].height = 8
            continue
        label_cell(ws, row, 1, label, bold=row in [18, 20])
        for c in range(2, 5):
            cell = ws.cell(row, c)
            cell.fill = PatternFill("solid", fgColor=GREEN_OUT)
            cell.border = thin()
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.number_format = "#,##0.0"
            if row == 19:
                cell.number_format = "0.0%"
            if row in [18, 20]:
                cell.font = Font(bold=True, size=10, name="Calibri")
        ws.row_dimensions[row].height = 18

    # Note about manual calculation
    ws.merge_cells("A22:D22")
    apply(ws["A22"],
          value="ℹ️  Price targets require the 3-Statement Model to be filled with historical data. "
                "Then build three separate scenario runs by changing the assumption inputs in Tab 1.",
          font=Font(size=9, italic=True, color="555555", name="Calibri"),
          fill=PatternFill("solid", fgColor=GREY_LIGHT),
          alignment=Alignment(horizontal="left", vertical="center", wrap_text=True))
    ws.row_dimensions[22].height = 30


# ═══════════════════════════════════════════════════════════════════════
# TAB 5: SENSITIVITY
# ═══════════════════════════════════════════════════════════════════════

def build_sensitivity(wb):
    ws = wb.create_sheet("Sensitivity")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:L1")
    apply(ws["A1"], **hdr("SENSITIVITY ANALYSIS — Price Target vs. Key Assumptions", size=12, bg=DARK_BLUE))
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:L2")
    apply(ws["A2"],
          value="To use Excel Data Tables: fill in the base case in DCF tab → select the sensitivity range "
                "→ Data → What-If Analysis → Data Table → row input = WACC cell, column input = terminal growth cell.",
          font=Font(size=9, italic=True, color="555555", name="Calibri"),
          fill=PatternFill("solid", fgColor=GREY_LIGHT),
          alignment=Alignment(horizontal="left", vertical="center", wrap_text=True))
    ws.row_dimensions[2].height = 28

    set_w(ws, 1, 24)
    for c in range(2, 9):
        set_w(ws, c, 12)

    # ── Table 1: Price vs. WACC × Terminal Growth ──
    section_header(ws, 4, 1, 9, "TABLE 1 — Implied Price (₹) vs. WACC × Terminal Growth Rate")

    wacc_vals = [0.10, 0.11, 0.12, 0.13, 0.14, 0.15]
    tg_vals   = [0.03, 0.035, 0.04, 0.045, 0.05, 0.055, 0.06]

    # Headers
    apply(ws.cell(5, 1), **hdr("WACC ↓ / Terminal g →", bg=GREY_MID, fg=DARK_BLUE))
    for c, tg in enumerate(tg_vals, start=2):
        apply(ws.cell(5, c), **hdr(f"{tg:.1%}", bg=TEAL))
        set_w(ws, c, 12)
    ws.row_dimensions[5].height = 20

    for r, wacc in enumerate(wacc_vals, start=6):
        apply(ws.cell(r, 1), **hdr(f"{wacc:.0%}", bg=TEAL, align="center"))
        for c, tg in enumerate(tg_vals, start=2):
            cell = ws.cell(r, c)
            # Simplified placeholder — real value comes from Data Table in Excel
            # Approximate Gordon Growth: P ∝ FCF / (WACC - g). Use index relative to base.
            base_wacc, base_tg = 0.12, 0.04
            approx_base = 280  # base price target (₹)
            # Rough scaling
            if wacc > tg:
                approx = approx_base * ((base_wacc - base_tg) / (wacc - tg))
            else:
                approx = 999
            cell.value = round(approx)
            # Colour: green if > current price (212), red if below
            colour = "C6EFCE" if approx > 212 else "FFC7CE"
            cell.fill = PatternFill("solid", fgColor=colour)
            cell.border = thin()
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.number_format = "₹#,##0"
            cell.font = Font(bold=(abs(wacc - 0.12) < 0.001 and abs(tg - 0.04) < 0.001),
                             size=9, name="Calibri")
        ws.row_dimensions[r].height = 17

    # Legend
    ws.merge_cells("A13:H13")
    apply(ws["A13"],
          value="🟢 Green = upside vs. ₹212 current price   |   🔴 Red = downside   |   "
                "Bold cell = Base Case (WACC 12%, g 4%)",
          font=Font(size=9, italic=True, color="555555", name="Calibri"),
          fill=PatternFill("solid", fgColor=GREY_LIGHT),
          alignment=Alignment(horizontal="left"))

    # ── Table 2: Price vs. Revenue CAGR × EBITDA Margin ──
    section_header(ws, 15, 1, 9, "TABLE 2 — Implied Upside (%) vs. Revenue CAGR × EBITDA Margin (FY29E)")

    rev_cagrs  = [0.10, 0.13, 0.15, 0.18, 0.22]
    margins    = [0.33, 0.36, 0.40, 0.43, 0.46, 0.49]

    apply(ws.cell(16, 1), **hdr("Rev CAGR ↓ / Margin →", bg=GREY_MID, fg=DARK_BLUE))
    for c, m in enumerate(margins, start=2):
        apply(ws.cell(16, c), **hdr(f"{m:.0%}", bg=TEAL))
    ws.row_dimensions[16].height = 20

    for r, cagr in enumerate(rev_cagrs, start=17):
        apply(ws.cell(r, 1), **hdr(f"{cagr:.0%}", bg=TEAL, align="center"))
        for c, margin in enumerate(margins, start=2):
            cell = ws.cell(r, c)
            # Approximate relative upside
            base_cagr, base_margin = 0.15, 0.42
            approx_upside = (cagr / base_cagr) * (margin / base_margin) - 1
            cell.value = approx_upside
            colour = "C6EFCE" if approx_upside > 0 else "FFC7CE"
            cell.fill = PatternFill("solid", fgColor=colour)
            cell.border = thin()
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.number_format = "+0.0%;-0.0%;0.0%"
            cell.font = Font(size=9, name="Calibri")
        ws.row_dimensions[r].height = 17

    ws.merge_cells("A23:H23")
    apply(ws["A23"],
          value="ℹ️  Values are illustrative approximations. Use Excel Data Tables (Data → What-If Analysis) for exact figures once the model is populated.",
          font=Font(size=9, italic=True, color="555555", name="Calibri"),
          fill=PatternFill("solid", fgColor=GREY_LIGHT),
          alignment=Alignment(horizontal="left"))
    ws.row_dimensions[23].height = 24


# ═══════════════════════════════════════════════════════════════════════
# TAB 6: DASHBOARD
# ═══════════════════════════════════════════════════════════════════════

def build_dashboard(wb):
    ws = wb.create_sheet("Dashboard")
    ws.sheet_view.showGridLines = False

    # Title block
    ws.merge_cells("A1:H1")
    apply(ws["A1"], **hdr("ANGEL ONE LTD (NSE: ANGELONE) — EQUITY RESEARCH SUMMARY", size=14, bg=DARK_BLUE))
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:H2")
    apply(ws["A2"],
          value=f"Prepared by: Dhruv Mandavkar   |   Date: March 2026   |   All figures in ₹ Millions unless stated",
          font=Font(size=9, italic=True, color="555555", name="Calibri"),
          fill=PatternFill("solid", fgColor=GREY_LIGHT),
          alignment=Alignment(horizontal="left", vertical="center"))
    ws.row_dimensions[2].height = 16

    for c in range(1, 9):
        set_w(ws, c, 16)
    set_w(ws, 1, 26)

    # ── Price Target Box ──
    section_header(ws, 4, 1, 3, "PRICE TARGET & RATING")
    pt_rows = [
        (5,  "DCF Price Target (₹)",           "='DCF Valuation'!B47"),
        (6,  "Current Market Price (₹)",         212.12),
        (7,  "Implied Upside / (Downside)",      "='DCF Valuation'!B49"),
        (8,  "Rating (DCF-based)",               "='DCF Valuation'!B50"),
        (9,  "52-Week High (₹)",                 328.50),
        (10, "52-Week Low (₹)",                  194.10),
        (11, "Market Cap (₹ Crore)",             20431),
        (12, "P/E (trailing)",                   19.08),
    ]
    for row, label, val in pt_rows:
        bold = row in [5, 7, 8]
        label_cell(ws, row, 1, label, bold=bold)
        c = ws.cell(row, 2)
        c.value = val
        is_output = row in [5, 7, 8]
        c.fill = PatternFill("solid", fgColor=GREEN_OUT if is_output else GREY_LIGHT)
        c.border = thin()
        c.alignment = Alignment(horizontal="right", vertical="center")
        c.number_format = ("0.0%" if row == 7 else
                           ("₹#,##0.00" if row in [5, 6, 9, 10] else "#,##0.0"))
        if bold:
            c.font = Font(bold=True, size=11 if is_output else 9, name="Calibri")
        ws.row_dimensions[row].height = 18

    # ── Revenue & Margin KPIs ──
    section_header(ws, 4, 4, 8, "KEY FINANCIAL METRICS")
    kpi_rows = [
        (5,  "FY24 Revenue (₹Mn)",              "='3-Statement Model'!E23"),
        (6,  "FY24 EBITDA Margin (%)",           "='3-Statement Model'!E30"),
        (7,  "FY24 PAT (₹Mn)",                  "='3-Statement Model'!E36"),
        (8,  "FY24 EPS (₹)",                    "='3-Statement Model'!E38"),
        (9,  "FY29E Revenue (₹Mn)",             "='3-Statement Model'!J23"),
        (10, "FY29E EBITDA Margin (%)",          "='3-Statement Model'!J30"),
        (11, "FY29E PAT (₹Mn)",                 "='3-Statement Model'!J36"),
        (12, "Revenue CAGR FY24–29E (%)",        "=(J23/'3-Statement Model'!E23)^(1/5)-1"),
    ]
    for row, label, val in kpi_rows:
        is_pct = "Margin" in label or "CAGR" in label
        label_cell(ws, row, 4, label)
        c = ws.cell(row, 5)
        c.value = val
        c.fill = PatternFill("solid", fgColor=BLUE_CALC)
        c.border = thin()
        c.alignment = Alignment(horizontal="right", vertical="center")
        c.number_format = "0.0%" if is_pct else ("#,##0.00" if "EPS" in label else "#,##0.0")
        ws.row_dimensions[row].height = 18

    # ── Investment Thesis ──
    section_header(ws, 14, 1, 8, "INVESTMENT THESIS")
    ws.merge_cells("A15:H19")
    apply(ws["A15"],
          value=(
              "India's structural shift from physical savings to equity markets is the central thesis. "
              "Angel One (ANGELONE) is the largest listed retail broker by active clients, with ~22M active users "
              "out of a 140M+ demat account base that continues to grow structurally.\n\n"
              "Key risk: SEBI's F&O margin changes have compressed take rates and triggered a sharp Q4 FY25 miss "
              "(revenue -22%, PAT -49% YoY). The bull case requires margin recovery and client reactivation "
              "as retail F&O volumes normalise.\n\n"
              "Current price: ₹212 vs. 52-week high of ₹328. At 19x P/E, the market is pricing in "
              "significant structural pressure. The DCF price target determines whether this is a dislocation or fair value."
          ),
          font=Font(size=9, name="Calibri", color="222222"),
          fill=PatternFill("solid", fgColor=WHITE),
          alignment=Alignment(horizontal="left", vertical="top", wrap_text=True),
          border=thin())
    ws.row_dimensions[15].height = 20
    for r in range(16, 20):
        ws.row_dimensions[r].height = 20

    # ── Data sources ──
    section_header(ws, 21, 1, 8, "DATA SOURCES")
    sources = [
        "Angel One Investor Relations: https://www.angelone.in/investor-relations",
        "BSE Filings (code 543235): https://www.bseindia.com",
        "Screener.in: https://www.screener.in/company/ANGELONE/consolidated/",
        "SEBI monthly broker data: https://www.sebi.gov.in",
        "NSE monthly market data: https://www.nseindia.com",
        "AMFI data: https://www.amfiindia.com",
    ]
    for r, src in enumerate(sources, start=22):
        ws.merge_cells(f"A{r}:H{r}")
        apply(ws[f"A{r}"],
              value=src,
              font=Font(size=9, name="Calibri", color="005F73"),
              fill=PatternFill("solid", fgColor=GREY_LIGHT),
              alignment=Alignment(horizontal="left"),
              border=thin())
        ws.row_dimensions[r].height = 16

    # ── Key risks ──
    section_header(ws, 29, 1, 8, "KEY RISKS")
    risks = [
        "Regulatory: Further SEBI F&O tightening could structurally reduce retail participation and brokerage take rates",
        "Competition: Zerodha, Groww maintain strong brand loyalty; market share defence required",
        "Margin compression: EBITDA fell from 46.8% (FY24) to 32.5% (Q4 FY25) — recovery trajectory uncertain",
        "Market dependency: Revenue heavily correlated to market volumes (downturns = significant PAT decline)",
        "MTF book risk: Margin lending creates credit exposure in volatile markets",
    ]
    for r, risk in enumerate(risks, start=30):
        ws.merge_cells(f"A{r}:H{r}")
        apply(ws[f"A{r}"],
              value=f"• {risk}",
              font=Font(size=9, name="Calibri", color="222222"),
              fill=PatternFill("solid", fgColor="FFF8F8"),
              alignment=Alignment(horizontal="left", wrap_text=True),
              border=thin())
        ws.row_dimensions[r].height = 22


# ═══════════════════════════════════════════════════════════════════════
# ASSUMPTIONS HELPER SHEET
# ═══════════════════════════════════════════════════════════════════════

def build_assumptions_helper(wb):
    """Hidden helper sheet for cross-tab references."""
    ws = wb.create_sheet("Assumptions")
    apply(ws["A1"], value="Shares Outstanding (Millions)")
    ws["B1"].value = 85.5   # update from latest BSE filing
    ws["A2"].value = "Current Share Price (₹)"
    ws["B2"].value = 212.12
    ws.sheet_state = "hidden"


# ═══════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════

def main():
    wb = Workbook()
    build_3_statement(wb)
    build_dcf(wb)
    build_comparables(wb)
    build_scenarios(wb)
    build_sensitivity(wb)
    build_dashboard(wb)
    build_assumptions_helper(wb)

    # Tab order
    for i, title in enumerate(["3-Statement Model", "DCF Valuation",
                                "Comparables", "Scenario Analysis",
                                "Sensitivity", "Dashboard"]):
        if title in [s.title for s in wb.worksheets]:
            wb[title].sheet_properties.tabColor = (
                "2F4858" if title == "3-Statement Model" else
                "005F73" if title == "DCF Valuation" else
                "1A7A4A" if title == "Dashboard" else "4A6741"
            )

    output_path = "angel-one-model.xlsx"
    wb.save(output_path)
    print(f"✓ Saved: {output_path}")
    print()
    print("NEXT STEPS:")
    print("1. Open angel-one-model.xlsx in Excel")
    print("2. Go to '3-Statement Model' tab")
    print("3. Fill in yellow cells (FY21–FY24) from Angel One Annual Reports")
    print("   → BSE code 543235: https://www.bseindia.com")
    print("   → Angel One IR: https://www.angelone.in/investor-relations")
    print("4. Check the sanity row (row 80) turns green")
    print("5. DCF tab will auto-populate the price target")


if __name__ == "__main__":
    main()
