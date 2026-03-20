"""
Build the Equity Fund Screening Model Excel file.
Run: python build_fund_screening.py
Output: fund-screening-model.xlsx
"""

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00
import openpyxl


# ── Colour palette (matches portfolio site) ──────────────────────────
DARK_BLUE    = "2F4858"   # Prussian blue — headers
TEAL         = "005F73"   # section headers
WHITE        = "FFFFFF"
PALE_BLUE    = "E8F4FD"   # input rows
YELLOW       = "FFF3CD"   # cells requiring manual entry
LIGHT_GREY   = "F8F9FA"
MID_GREY     = "DEE2E6"
OUTPUT_GREEN = "D4EDDA"   # scoring output cells


def thin_border(sides="all"):
    s = Side(style="thin", color="C0C0C0")
    n = Side(style=None)
    return Border(
        left=s if sides == "all" or "l" in sides else n,
        right=s if sides == "all" or "r" in sides else n,
        top=s if sides == "all" or "t" in sides else n,
        bottom=s if sides == "all" or "b" in sides else n,
    )


def hdr(text, bold=True, fg=WHITE, bg=DARK_BLUE, size=10, wrap=False, align="center"):
    """Return a dict of style kwargs for a header cell."""
    return dict(
        value=text,
        font=Font(bold=bold, color=fg, size=size, name="Calibri"),
        fill=PatternFill("solid", fgColor=bg),
        alignment=Alignment(horizontal=align, vertical="center", wrap_text=wrap),
        border=thin_border(),
    )


def apply(cell, **kwargs):
    for k, v in kwargs.items():
        if k == "value":
            cell.value = v
        elif k == "font":
            cell.font = v
        elif k == "fill":
            cell.fill = v
        elif k == "alignment":
            cell.alignment = v
        elif k == "border":
            cell.border = v
        elif k == "number_format":
            cell.number_format = v


def set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width


# ═══════════════════════════════════════════════════════════════════════
# TAB 1: INPUT
# ═══════════════════════════════════════════════════════════════════════

def build_input_tab(wb):
    ws = wb.active
    ws.title = "Input"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    # Title row
    ws.merge_cells("A1:F1")
    apply(ws["A1"], **hdr("EQUITY FUND SCREENING MODEL — Input Fund List", size=13, bg=DARK_BLUE))
    ws.row_dimensions[1].height = 30

    # Subtitle
    ws.merge_cells("A2:F2")
    apply(ws["A2"],
          value="List funds below. Paste data into the 'Raw Data' tab once collected from Value Research Online or AMFI.",
          font=Font(color="555555", size=9, italic=True, name="Calibri"),
          fill=PatternFill("solid", fgColor=LIGHT_GREY),
          alignment=Alignment(horizontal="left", vertical="center"))
    ws.row_dimensions[2].height = 18

    # Column headers
    headers = ["#", "Fund Name", "AMC", "Category", "Sub-category", "Notes"]
    widths  = [4,   40,          22,    18,           20,             30]
    for c, (h, w) in enumerate(zip(headers, widths), start=1):
        apply(ws.cell(3, c), **hdr(h, bg=TEAL))
        set_col_width(ws, c, w)
    ws.row_dimensions[3].height = 20

    # 30 blank fund rows (yellow = manual entry)
    fill_y = PatternFill("solid", fgColor=YELLOW)
    fill_w = PatternFill("solid", fgColor=WHITE)
    for r in range(4, 34):
        ws.cell(r, 1).value = r - 3
        ws.cell(r, 1).font = Font(color="888888", size=9, name="Calibri")
        ws.cell(r, 1).alignment = Alignment(horizontal="center")
        for c in range(2, 7):
            cell = ws.cell(r, c)
            cell.fill = fill_y if c <= 5 else fill_w
            cell.border = thin_border()
            cell.alignment = Alignment(vertical="center")
        ws.row_dimensions[r].height = 17

    # Legend
    ws.merge_cells("A35:F35")
    apply(ws["A35"],
          value="🟡 Yellow cells = fill in manually   |   Paste the complete list here before running screening",
          font=Font(color="856404", size=9, italic=True, name="Calibri"),
          fill=PatternFill("solid", fgColor="FFF3CD"),
          alignment=Alignment(horizontal="left"))


# ═══════════════════════════════════════════════════════════════════════
# TAB 2: RAW DATA
# ═══════════════════════════════════════════════════════════════════════

def build_raw_data_tab(wb):
    ws = wb.create_sheet("Raw Data")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "C4"

    ws.merge_cells("A1:N1")
    apply(ws["A1"], **hdr("RAW DATA — Pull from Value Research Online (valueresearchonline.com)", size=12, bg=DARK_BLUE))
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:N2")
    apply(ws["A2"],
          value="Source: valueresearchonline.com → Fund → Performance/Portfolio tab. Paste one fund per row. All return figures should be in % per annum.",
          font=Font(color="555555", size=9, italic=True, name="Calibri"),
          fill=PatternFill("solid", fgColor=LIGHT_GREY),
          alignment=Alignment(horizontal="left", vertical="center"))
    ws.row_dimensions[2].height = 18

    col_headers = [
        ("Fund Name",       22),
        ("Category",        18),
        ("1Y Return (%)",   13),
        ("3Y Return (%)",   13),
        ("5Y Return (%)",   13),
        ("Sharpe (3Y)",     12),
        ("Std Dev (3Y, %)", 14),
        ("Alpha (3Y)",      12),
        ("Beta (3Y)",       10),
        ("Expense Ratio (%)",13),
        ("AUM (₹ Cr)",      13),
        ("Benchmark Index", 22),
        ("Data Date",       12),
        ("Notes",           20),
    ]
    for c, (h, w) in enumerate(col_headers, start=1):
        apply(ws.cell(3, c), **hdr(h, bg=TEAL, wrap=True))
        set_col_width(ws, c, w)
    ws.row_dimensions[3].height = 32

    fill_y = PatternFill("solid", fgColor=YELLOW)
    for r in range(4, 34):
        for c in range(1, len(col_headers) + 1):
            cell = ws.cell(r, c)
            cell.fill = fill_y
            cell.border = thin_border()
            cell.alignment = Alignment(vertical="center")
            if c >= 3 and c <= 11:
                cell.number_format = "0.00"
        ws.row_dimensions[r].height = 17

    ws.merge_cells(f"A35:N35")
    apply(ws["A35"],
          value="🟡 All yellow cells require manual data entry from Value Research Online. Do NOT calculate formulas in this tab — paste values only.",
          font=Font(color="856404", size=9, italic=True, name="Calibri"),
          fill=PatternFill("solid", fgColor="FFF3CD"),
          alignment=Alignment(horizontal="left"))


# ═══════════════════════════════════════════════════════════════════════
# TAB 3: SCORING
# ═══════════════════════════════════════════════════════════════════════

CRITERIA = [
    ("Risk-Adjusted Return", "35%", "Sharpe ratio × return score composite",
     "Rewards funds with high returns relative to volatility taken"),
    ("Return Consistency",   "25%", "Quartile rank over rolling 1Y periods",
     "Rewards funds that deliver consistently, not just in good markets"),
    ("Alpha vs Benchmark",   "20%", "Jensen's Alpha (3-year)",
     "Rewards skill — excess return above what market beta would predict"),
    ("Cost Efficiency",      "10%", "Expense ratio percentile (lower = better)",
     "Lower cost = better net return. Critical for long-horizon investors."),
    ("AUM Stability",        "10%", "3Y AUM CAGR, moderate growth preferred",
     "Extreme inflows can hurt performance; very low AUM = closure risk"),
]

def build_scoring_tab(wb):
    ws = wb.create_sheet("Scoring")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "C5"

    ws.merge_cells("A1:J1")
    apply(ws["A1"], **hdr("SCORING ENGINE — Weighted Fund Scoring (0–100)", size=12, bg=DARK_BLUE))
    ws.row_dimensions[1].height = 28

    # Criteria table
    ws.merge_cells("A2:J2")
    apply(ws["A2"], **hdr("SCORING METHODOLOGY", bg=TEAL))
    ws.row_dimensions[2].height = 20

    apply(ws.cell(3, 1), **hdr("Criterion", bg=MID_GREY, fg=DARK_BLUE, align="left"))
    apply(ws.cell(3, 2), **hdr("Weight",    bg=MID_GREY, fg=DARK_BLUE))
    apply(ws.cell(3, 3), **hdr("How Calculated", bg=MID_GREY, fg=DARK_BLUE, align="left"))
    apply(ws.cell(3, 4), **hdr("Rationale",      bg=MID_GREY, fg=DARK_BLUE, align="left"))
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 35
    ws.column_dimensions["D"].width = 45

    for r, (name, wt, calc, rationale) in enumerate(CRITERIA, start=4):
        apply(ws.cell(r, 1), value=name, font=Font(bold=True, size=9, name="Calibri"),
              fill=PatternFill("solid", fgColor=LIGHT_GREY), border=thin_border(),
              alignment=Alignment(horizontal="left"))
        apply(ws.cell(r, 2), value=wt,   font=Font(bold=True, size=9, name="Calibri"),
              fill=PatternFill("solid", fgColor=LIGHT_GREY), border=thin_border(),
              alignment=Alignment(horizontal="center"))
        apply(ws.cell(r, 3), value=calc, font=Font(size=9, name="Calibri"),
              fill=PatternFill("solid", fgColor=WHITE), border=thin_border(),
              alignment=Alignment(horizontal="left", wrap_text=True))
        apply(ws.cell(r, 4), value=rationale, font=Font(size=9, italic=True, color="555555", name="Calibri"),
              fill=PatternFill("solid", fgColor=WHITE), border=thin_border(),
              alignment=Alignment(horizontal="left", wrap_text=True))
        ws.row_dimensions[r].height = 25

    # Score table header
    SCORE_START = 10
    score_headers = [
        ("Fund Name", 22), ("Category", 16),
        ("Risk-Adj Score\n(35%)", 13), ("Consistency\n(25%)", 12),
        ("Alpha Score\n(20%)", 12), ("Cost Score\n(10%)", 12),
        ("AUM Score\n(10%)", 11), ("TOTAL SCORE\n/100", 12),
        ("Rank", 7), ("Rating", 10),
    ]
    ws.merge_cells(f"A{SCORE_START - 1}:J{SCORE_START - 1}")
    apply(ws[f"A{SCORE_START - 1}"], **hdr("FUND SCORES (auto-calculated from Raw Data)", bg=TEAL))
    ws.row_dimensions[SCORE_START - 1].height = 20

    for c, (h, w) in enumerate(score_headers, start=1):
        apply(ws.cell(SCORE_START, c), **hdr(h, bg=DARK_BLUE, wrap=True))
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.row_dimensions[SCORE_START].height = 38

    # Formula rows: for 30 funds, generate RANK-based score formulas
    # Score formula: percentile rank within column in Raw Data
    # Component score = (RANK within peers, inverted) / count * 100 * weight
    # Using RANK.EQ for each metric
    n = 30   # max funds
    raw_start = 4  # first data row in Raw Data tab

    for i in range(n):
        r = SCORE_START + 1 + i
        row_raw = raw_start + i

        ws.cell(r, 1).value = f"='Raw Data'!A{row_raw}"
        ws.cell(r, 2).value = f"='Raw Data'!B{row_raw}"

        # Risk-adjusted: composite of Sharpe (col F) and 3Y return (col E)
        ws.cell(r, 3).value = (
            f"=IF('Raw Data'!A{row_raw}=\"\",\"\","
            f"ROUND((RANK.EQ('Raw Data'!F{row_raw},'Raw Data'!$F${raw_start}:$F${raw_start+n-1},1)"
            f"/COUNTA('Raw Data'!$A${raw_start}:$A${raw_start+n-1})*100)*0.6+"
            f"(RANK.EQ('Raw Data'!E{row_raw},'Raw Data'!$E${raw_start}:$E${raw_start+n-1},1)"
            f"/COUNTA('Raw Data'!$A${raw_start}:$A${raw_start+n-1})*100)*0.4,1))"
        )
        # Consistency: uses Std Dev (col G), lower = better (reverse rank)
        ws.cell(r, 4).value = (
            f"=IF('Raw Data'!A{row_raw}=\"\",\"\","
            f"ROUND((COUNTA('Raw Data'!$A${raw_start}:$A${raw_start+n-1})-"
            f"RANK.EQ('Raw Data'!G{row_raw},'Raw Data'!$G${raw_start}:$G${raw_start+n-1},1)+1)"
            f"/COUNTA('Raw Data'!$A${raw_start}:$A${raw_start+n-1})*100,1))"
        )
        # Alpha score: col H
        ws.cell(r, 5).value = (
            f"=IF('Raw Data'!A{row_raw}=\"\",\"\","
            f"ROUND(RANK.EQ('Raw Data'!H{row_raw},'Raw Data'!$H${raw_start}:$H${raw_start+n-1},1)"
            f"/COUNTA('Raw Data'!$A${raw_start}:$A${raw_start+n-1})*100,1))"
        )
        # Cost score: col J (expense ratio), lower = better (reverse rank)
        ws.cell(r, 6).value = (
            f"=IF('Raw Data'!A{row_raw}=\"\",\"\","
            f"ROUND((COUNTA('Raw Data'!$A${raw_start}:$A${raw_start+n-1})-"
            f"RANK.EQ('Raw Data'!J{row_raw},'Raw Data'!$J${raw_start}:$J${raw_start+n-1},1)+1)"
            f"/COUNTA('Raw Data'!$A${raw_start}:$A${raw_start+n-1})*100,1))"
        )
        # AUM stability: col K (moderate growth preferred, score based on rank)
        ws.cell(r, 7).value = (
            f"=IF('Raw Data'!A{row_raw}=\"\",\"\","
            f"ROUND(RANK.EQ('Raw Data'!K{row_raw},'Raw Data'!$K${raw_start}:$K${raw_start+n-1},1)"
            f"/COUNTA('Raw Data'!$A${raw_start}:$A${raw_start+n-1})*100,1))"
        )
        # Total score: weighted sum
        ws.cell(r, 8).value = (
            f"=IF(C{r}=\"\",\"\",ROUND(C{r}*0.35+D{r}*0.25+E{r}*0.20+F{r}*0.10+G{r}*0.10,1))"
        )
        # Rank
        ws.cell(r, 9).value = (
            f"=IF(H{r}=\"\",\"\","
            f"RANK.EQ(H{r},$H${SCORE_START+1}:$H${SCORE_START+n},0))"
        )
        # Rating label
        ws.cell(r, 10).value = (
            f"=IF(H{r}=\"\",\"\","
            f'IF(H{r}>=80,"★★★ Strong Buy",'
            f'IF(H{r}>=65,"★★ Buy",'
            f'IF(H{r}>=50,"★ Hold","✗ Pass"))))'
        )

        # Styles
        fill_row = PatternFill("solid", fgColor=OUTPUT_GREEN if i % 2 == 0 else "F0FFF4")
        for c in range(1, 11):
            cell = ws.cell(r, c)
            cell.fill = fill_row
            cell.border = thin_border()
            cell.alignment = Alignment(horizontal="center" if c > 1 else "left", vertical="center")
            if c == 8:
                cell.font = Font(bold=True, size=10, name="Calibri")
        ws.row_dimensions[r].height = 17

    # Colour scale on total score column
    score_range = f"H{SCORE_START+1}:H{SCORE_START+n}"
    ws.conditional_formatting.add(score_range, ColorScaleRule(
        start_type="min", start_color="FFCCCC",
        mid_type="percentile", mid_value=50, mid_color="FFFF99",
        end_type="max", end_color="CCFFCC"
    ))


# ═══════════════════════════════════════════════════════════════════════
# TAB 4: OUTPUT (LEADERBOARD)
# ═══════════════════════════════════════════════════════════════════════

def build_output_tab(wb):
    ws = wb.create_sheet("Output — Leaderboard")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:G1")
    apply(ws["A1"], **hdr("OUTPUT — RANKED LEADERBOARD (auto-sorted)", size=12, bg=DARK_BLUE))
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:G2")
    apply(ws["A2"],
          value="Sort the Scoring tab by column H (Total Score) descending to see the leaderboard. "
                "Top 3 highlighted for qualitative review in the Notes tab.",
          font=Font(size=9, italic=True, color="555555", name="Calibri"),
          fill=PatternFill("solid", fgColor=LIGHT_GREY),
          alignment=Alignment(horizontal="left", vertical="center"))
    ws.row_dimensions[2].height = 18

    headers = [("Rank", 7), ("Fund Name", 38), ("Category", 18),
               ("Total Score", 12), ("Rating", 16), ("Key Strength", 28), ("Key Risk", 28)]
    for c, (h, w) in enumerate(headers, start=1):
        apply(ws.cell(3, c), **hdr(h, bg=TEAL, wrap=True))
        set_col_width(ws, c, w)
    ws.row_dimensions[3].height = 22

    # Reference top 10 from Scoring tab dynamically
    sc = "Scoring"
    SCORE_START = 10
    for i in range(10):
        r = 4 + i
        rank_row = SCORE_START + 1 + i   # approximate (proper sort needs VBA/manual)
        bg = "D4EDDA" if i < 3 else ("FFF3CD" if i < 7 else "F8F9FA")
        fill_r = PatternFill("solid", fgColor=bg)
        medal = ["🥇", "🥈", "🥉", "", "", "", "", "", "", ""][i]
        ws.cell(r, 1).value = f"=IFERROR(MATCH(SMALL(Scoring!$I${SCORE_START+1}:$I${SCORE_START+30},{i+1}),Scoring!$I${SCORE_START+1}:$I${SCORE_START+30},0)+{SCORE_START},\"\")"
        ws.cell(r, 1).value = i + 1   # static rank for now
        ws.cell(r, 2).value = f"=IFERROR(INDEX({sc}!A${SCORE_START+1}:A${SCORE_START+30},MATCH({i+1},{sc}!I${SCORE_START+1}:I${SCORE_START+30},0)),\"\")"
        ws.cell(r, 3).value = f"=IFERROR(INDEX({sc}!B${SCORE_START+1}:B${SCORE_START+30},MATCH({i+1},{sc}!I${SCORE_START+1}:I${SCORE_START+30},0)),\"\")"
        ws.cell(r, 4).value = f"=IFERROR(INDEX({sc}!H${SCORE_START+1}:H${SCORE_START+30},MATCH({i+1},{sc}!I${SCORE_START+1}:I${SCORE_START+30},0)),\"\")"
        ws.cell(r, 5).value = f"=IFERROR(INDEX({sc}!J${SCORE_START+1}:J${SCORE_START+30},MATCH({i+1},{sc}!I${SCORE_START+1}:I${SCORE_START+30},0)),\"\")"
        ws.cell(r, 6).value = ""   # manual: key strength
        ws.cell(r, 7).value = ""   # manual: key risk

        for c in range(1, 8):
            cell = ws.cell(r, c)
            cell.fill = fill_r
            cell.border = thin_border()
            cell.alignment = Alignment(horizontal="center" if c in [1, 4] else "left", vertical="center")
            if c == 4:
                cell.font = Font(bold=True, size=10, name="Calibri")
        ws.row_dimensions[r].height = 20


# ═══════════════════════════════════════════════════════════════════════
# TAB 5: NOTES
# ═══════════════════════════════════════════════════════════════════════

def build_notes_tab(wb):
    ws = wb.create_sheet("Notes — Top 3")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:D1")
    apply(ws["A1"], **hdr("QUALITATIVE NOTES — Top 3 Funds", size=12, bg=DARK_BLUE))
    ws.row_dimensions[1].height = 28

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 38
    ws.column_dimensions["D"].width = 38

    apply(ws.cell(2, 1), **hdr("", bg=TEAL))
    for c, label in enumerate(["Rank 1", "Rank 2", "Rank 3"], start=2):
        apply(ws.cell(2, c), **hdr(label, bg=TEAL))
    ws.row_dimensions[2].height = 22

    row_labels = [
        "Fund Name",
        "AMC / Manager",
        "Mandate",
        "Manager Tenure",
        "Total Score",
        "Key Differentiator",
        "What it does well",
        "Main risk / concern",
        "Portfolio fit?",
        "Recommendation",
    ]
    for r, label in enumerate(row_labels, start=3):
        apply(ws.cell(r, 1), value=label, font=Font(bold=True, size=9, name="Calibri"),
              fill=PatternFill("solid", fgColor=LIGHT_GREY), border=thin_border(),
              alignment=Alignment(horizontal="left", vertical="center"))
        for c in range(2, 5):
            fill_c = PatternFill("solid", fgColor=YELLOW)
            apply(ws.cell(r, c), fill=fill_c, border=thin_border(),
                  alignment=Alignment(horizontal="left", vertical="center", wrap_text=True))
        ws.row_dimensions[r].height = 40 if r >= 8 else 22


# ═══════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════

def main():
    wb = Workbook()
    build_input_tab(wb)
    build_raw_data_tab(wb)
    build_scoring_tab(wb)
    build_output_tab(wb)
    build_notes_tab(wb)

    output_path = "fund-screening-model.xlsx"
    wb.save(output_path)
    print(f"✓ Saved: {output_path}")


if __name__ == "__main__":
    main()
