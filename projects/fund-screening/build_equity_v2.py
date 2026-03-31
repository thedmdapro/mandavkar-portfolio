"""
build_equity_v2.py — Equity Fund Screening Model v2
Phase 1: 8 SEBI equity categories from ACE MF Regular Growth data.

Categories scored: Large Cap, Flexi Cap, Mid Cap, Small Cap,
                   ELSS, Multi Cap, Large & Mid Cap, Focused Fund
Informational only: Dividend Yield, Contra

Run: python3 build_equity_v2.py
Output: fund-screening-model.xlsx
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
import warnings
warnings.filterwarnings('ignore')
from datetime import date

# ══════════════════════════════════════════════════════════════════════
# CONFIGURATION
# ══════════════════════════════════════════════════════════════════════

DATA_DATE = "March 2026"
TODAY     = date.today().strftime("%d %b %Y")

BASE = "/sessions/optimistic-zealous-ptolemy/mnt/Portfolio Website/projects/fund-screening/Ace MF Data Sheets/"
OUT  = "/sessions/optimistic-zealous-ptolemy/mnt/Portfolio Website/projects/fund-screening/fund-screening-model.xlsx"

# TRW institutional colour palette
NAVY       = "1F4E79"
LT_BLUE    = "D6E4F0"
ZEBRA      = "F5F8FB"
WHITE      = "FFFFFF"
CREAM      = "F5F8FB"
MID_GREY   = "DEE2E6"
DARK_TEXT  = "1A1A2E"
GOLD       = "B8860B"

# Rating band colours (fill / text)
BAND_STYLE = {
    "Strong Buy":  ("1a6b3c", "FFFFFF"),
    "Buy":         ("2e7d32", "FFFFFF"),
    "Hold":        ("b8860b", "FFFFFF"),
    "Underweight": ("c0392b", "FFFFFF"),
    "Avoid":       ("7b0000", "FFFFFF"),
    "—":           ("888888", "FFFFFF"),
}

# Scoring weights per category
WEIGHTS = {
    "default": {
        "alpha": 0.25, "sharpe": 0.25,
        "consistency": 0.25, "sortino": 0.20, "cost": 0.05
    },
    "Equity Linked Savings Scheme": {
        "alpha": 0.20, "sharpe": 0.25,
        "consistency": 0.30, "sortino": 0.20, "cost": 0.05
    },
}

# AUM gate per category (₹ Cr)
AUM_GATE = {
    "default": 500,
    "Equity Linked Savings Scheme": 300,
}

SCORED_CATS = [
    "Large Cap Fund", "Flexi Cap Fund", "Mid Cap Fund", "Small cap Fund",
    "Equity Linked Savings Scheme", "Multi Cap Fund",
    "Large & Mid Cap", "Focused Fund",
]
INFO_CATS = ["Dividend Yield", "Contra"]

TAB_NAMES = {
    "Large Cap Fund": "Large Cap",
    "Flexi Cap Fund": "Flexi Cap",
    "Mid Cap Fund":   "Mid Cap",
    "Small cap Fund": "Small Cap",
    "Equity Linked Savings Scheme": "ELSS",
    "Multi Cap Fund": "Multi Cap",
    "Large & Mid Cap": "Large & Mid Cap",
    "Focused Fund":   "Focused Fund",
    "Dividend Yield": "Div Yield (Info)",
    "Contra":         "Contra (Info)",
}

SEBI_MANDATES = {
    "Large Cap Fund":   "Min 80% in top-100 companies by market cap.",
    "Flexi Cap Fund":   "Min 65% equity; no market-cap restriction.",
    "Mid Cap Fund":     "Min 65% in companies ranked 101–250 by market cap.",
    "Small cap Fund":   "Min 65% in companies ranked 251+ by market cap.",
    "Equity Linked Savings Scheme": "Min 80% equity. Mandatory 3-year lock-in. Tax benefit u/s 80C.",
    "Multi Cap Fund":   "Min 25% each in large, mid, and small cap.",
    "Large & Mid Cap":  "Min 35% large cap + min 35% mid cap.",
    "Focused Fund":     "Max 30 stocks. Min 65% equity. High-conviction, concentrated portfolio.",
    "Dividend Yield":   "Min 65% in dividend-paying stocks.",
    "Contra":           "Contrarian investment strategy. Min 65% equity.",
}


# ══════════════════════════════════════════════════════════════════════
# STYLING HELPERS
# ══════════════════════════════════════════════════════════════════════

def thin_border(sides="all"):
    s = Side(style="thin", color="C0C0C0")
    n = Side(style=None)
    kw = {}
    for side in ["left", "right", "top", "bottom"]:
        kw[side] = s if (sides == "all" or side[0] in sides) else n
    return Border(**kw)

def apply(cell, value=None, bold=False, fg=DARK_TEXT, bg=None,
          size=10, wrap=False, halign="left", valign="center",
          border=True, italic=False, number_format=None):
    if value is not None:
        cell.value = value
    cell.font = Font(bold=bold, color=fg, size=size, name="Arial", italic=italic)
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=halign, vertical=valign, wrap_text=wrap)
    if border:
        cell.border = thin_border()
    if number_format:
        cell.number_format = number_format

def col_w(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

def row_h(ws, row, height):
    ws.row_dimensions[row].height = height

def hdr(ws, row, col, value, bg=NAVY, fg=WHITE, bold=True, size=10,
        halign="center", wrap=False, colspan=1):
    """Write a header cell, optionally merging columns."""
    cell = ws.cell(row, col)
    apply(cell, value=value, bold=bold, fg=fg, bg=bg, size=size,
          halign=halign, wrap=wrap)
    if colspan > 1:
        ws.merge_cells(
            start_row=row, start_column=col,
            end_row=row, end_column=col + colspan - 1
        )
    return cell


# ══════════════════════════════════════════════════════════════════════
# DATA PIPELINE
# ══════════════════════════════════════════════════════════════════════

def load_clean(filepath):
    """Load ACE MF export, filter to Regular Growth, dedup on primary benchmark."""
    df = pd.read_excel(filepath, header=1)
    df.columns = df.columns.str.strip()

    sn = "Scheme Name"
    # Regular Growth: contains (G), no IDCW, no Dir
    mask = (
        df[sn].str.contains(r'\(G\)', na=False) &
        ~df[sn].str.contains('IDCW', case=False, na=False) &
        ~df[sn].str.contains('Dir',  case=False, na=False)
    )
    df = df[mask].drop_duplicates(subset=[sn], keep='first').copy()

    # Numeric coerce key columns
    num_cols = [
        'AUMT_AUM', 'SD_Age (From Incept Date)',
        'SCAG_1YEAR_CAGR', 'SCAG_3YEAR_CAGR', 'SCAG_5YEAR_CAGR', 'SCAG_10YEAR_CAGR',
        'BMCAG_1YEAR_CAGR', 'BMCAG_3YEAR_CAGR', 'BMCAG_5YEAR_CAGR', 'BMCAG_10YEAR_CAGR',
        'RR3_Sharpe Ratio', 'RR3_Sortino', 'RR3_Std. Deviation',
        'RR3_Beta', 'RR3_Up Capture Ratio', 'RR3_Down Capture Ratio',
        'EXR_Expense Ratio(%)', 'EXR_Dir Plan Expense Ratio(%)',
        'RR3_Information Ratio',
    ]
    for c in num_cols:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors='coerce')

    return df


def compute_derived(df):
    """Add alpha columns and ER spread."""
    df = df.copy()
    for yr in ['1YEAR', '3YEAR', '5YEAR', '10YEAR']:
        f_col = f'SCAG_{yr}_CAGR'
        b_col = f'BMCAG_{yr}_CAGR'
        a_col = f'alpha_{yr}'
        if f_col in df.columns and b_col in df.columns:
            df[a_col] = df[f_col] - df[b_col]
        else:
            df[a_col] = np.nan

    er_reg = df.get('EXR_Expense Ratio(%)', pd.Series(np.nan, index=df.index))
    er_dir = df.get('EXR_Dir Plan Expense Ratio(%)', pd.Series(np.nan, index=df.index))
    df['er_spread'] = er_reg - er_dir
    return df


def apply_gates(df, category):
    """Return (passed_df, failed_df) based on gate criteria."""
    aum_thresh = AUM_GATE.get(category, AUM_GATE["default"])

    mask_aum   = df['AUMT_AUM'] >= aum_thresh
    mask_age   = df['SD_Age (From Incept Date)'] >= 3
    mask_alpha = (df['alpha_3YEAR'] > -5) | df['alpha_3YEAR'].isna()
    # If 3Y alpha is NaN (fund < 3Y), treat as failing the alpha gate too
    mask_alpha_strict = df['alpha_3YEAR'] > -5

    passed = df[mask_aum & mask_age & mask_alpha_strict].copy()
    failed = df[~(mask_aum & mask_age & mask_alpha_strict)].copy()
    return passed, failed


def percentile_rank(series, ascending=True):
    """Rank values as percentiles 0–100. ascending=True means higher value → higher rank."""
    n = series.notna().sum()
    if n == 0:
        return pd.Series(50.0, index=series.index)
    ranks = series.rank(method='average', na_option='keep', ascending=ascending)
    pct = (ranks / n) * 100
    return pct.fillna(50.0)


def score_category(df, category):
    """Compute 5-factor composite score (0–100) and assign rating band."""
    df = df.copy()
    w = WEIGHTS.get(category, WEIGHTS["default"])

    # Factor 1: Alpha (3Y) — higher is better
    df['pct_alpha'] = percentile_rank(df['alpha_3YEAR'], ascending=True)

    # Factor 2: Sharpe (3Y) — higher is better
    df['pct_sharpe'] = percentile_rank(df['RR3_Sharpe Ratio'], ascending=True)

    # Factor 3: Consistency — average percentile across available alpha horizons
    # Use 1Y, 3Y, 5Y alpha; exclude NaN periods
    alpha_cols = ['alpha_1YEAR', 'alpha_3YEAR', 'alpha_5YEAR']
    pct_parts = []
    for ac in alpha_cols:
        if ac in df.columns and df[ac].notna().sum() >= 3:
            pct_parts.append(percentile_rank(df[ac], ascending=True))
    if pct_parts:
        df['pct_consistency'] = pd.concat(pct_parts, axis=1).mean(axis=1)
    else:
        df['pct_consistency'] = 50.0

    # Factor 4: Sortino (3Y) — higher is better
    df['pct_sortino'] = percentile_rank(df['RR3_Sortino'], ascending=True)

    # Factor 5: Cost (ER) — lower is better → ascending=False
    df['pct_cost'] = percentile_rank(df['EXR_Expense Ratio(%)'], ascending=False)

    # Composite
    df['composite'] = (
        df['pct_alpha']       * w['alpha']       +
        df['pct_sharpe']      * w['sharpe']       +
        df['pct_consistency'] * w['consistency']  +
        df['pct_sortino']     * w['sortino']      +
        df['pct_cost']        * w['cost']
    ).round(1)

    # Rating bands
    def band(s):
        if pd.isna(s):   return "—"
        if s >= 75:      return "Strong Buy"
        elif s >= 60:    return "Buy"
        elif s >= 45:    return "Hold"
        elif s >= 30:    return "Underweight"
        else:            return "Avoid"

    df['rating'] = df['composite'].apply(band)
    df['rank']   = df['composite'].rank(method='min', ascending=False).astype(int)
    df = df.sort_values('rank')
    return df


# ══════════════════════════════════════════════════════════════════════
# EXCEL BUILDERS
# ══════════════════════════════════════════════════════════════════════

# --- COVER TAB ---
def build_cover(wb, all_stats):
    ws = wb.create_sheet("Cover")
    ws.sheet_view.showGridLines = False
    ws.sheet_view.showRowColHeaders = False

    # Title block
    ws.merge_cells("B2:I2")
    apply(ws["B2"], value="EQUITY FUND SCREENING MODEL", bold=True,
          fg=WHITE, bg=NAVY, size=18, halign="center")
    row_h(ws, 2, 36)

    ws.merge_cells("B3:I3")
    apply(ws["B3"], value="Systematic 3-Layer Screening | 8 SEBI Equity Categories | Regular Growth Plans",
          fg=WHITE, bg=NAVY, size=11, halign="center", italic=True)
    row_h(ws, 3, 22)

    ws.merge_cells("B4:I4")
    apply(ws["B4"], value=f"Data as of: {DATA_DATE}  |  Built: {TODAY}  |  v2.0",
          fg="AAAAAA", bg=NAVY, size=9, halign="center")
    row_h(ws, 4, 16)

    # Key stats
    row = 6
    ws.merge_cells(f"B{row}:I{row}")
    apply(ws.cell(row, 2), value="MODEL OVERVIEW", bold=True,
          fg=WHITE, bg=NAVY, size=12, halign="center")
    row_h(ws, row, 24)

    stats_data = [
        ("Starting Universe",  str(all_stats['total_raw'])),
        ("After Gate Filters", str(all_stats['total_scored'])),
        ("Categories Scored",  "8 SEBI Categories"),
        ("Scoring Factors",    "Alpha · Sharpe · Consistency · Sortino · Cost"),
        ("Rating Bands",       "Strong Buy / Buy / Hold / Underweight / Avoid"),
        ("Data Source",        "ACE Mutual Fund — Regular Growth Plans"),
    ]
    for i, (label, val) in enumerate(stats_data):
        r = row + 1 + i
        ws.merge_cells(f"B{r}:D{r}")
        apply(ws.cell(r, 2), value=label, bold=True, fg=DARK_TEXT, bg=LT_BLUE,
              size=10, halign="left")
        ws.merge_cells(f"E{r}:I{r}")
        apply(ws.cell(r, 5), value=val, fg=DARK_TEXT, bg=ZEBRA,
              size=10, halign="left")
        row_h(ws, r, 18)

    # Legend
    row = row + len(stats_data) + 2
    ws.merge_cells(f"B{row}:I{row}")
    apply(ws.cell(row, 2), value="RATING BAND LEGEND", bold=True,
          fg=WHITE, bg=NAVY, size=11, halign="center")
    row_h(ws, row, 22)

    legend = [
        ("Strong Buy",  "≥ 75", "Top quartile across all factors. Strong alpha, superior risk-adjusted returns."),
        ("Buy",         "60–74", "Above-average performance. Consistent outperformance vs benchmark."),
        ("Hold",        "45–59", "Mid-range. Watch for improvement or deterioration."),
        ("Underweight", "30–44", "Below-average. Underperforming on multiple factors."),
        ("Avoid",       "< 30",  "Bottom quartile. Poor alpha, high cost, or weak risk management."),
    ]
    for i, (rating, score_range, desc) in enumerate(legend):
        r = row + 1 + i
        fill_hex, text_hex = BAND_STYLE[rating]
        ws.merge_cells(f"B{r}:C{r}")
        apply(ws.cell(r, 2), value=rating, bold=True, fg=text_hex,
              bg=fill_hex, size=10, halign="center")
        apply(ws.cell(r, 4), value=score_range, fg=DARK_TEXT,
              bg=ZEBRA if i % 2 == 0 else WHITE, size=10, halign="center")
        ws.merge_cells(f"E{r}:I{r}")
        apply(ws.cell(r, 5), value=desc, fg=DARK_TEXT,
              bg=ZEBRA if i % 2 == 0 else WHITE, size=10, halign="left")
        row_h(ws, r, 18)

    # 3-layer framework
    row = row + len(legend) + 2
    ws.merge_cells(f"B{row}:I{row}")
    apply(ws.cell(row, 2), value="SCREENING METHODOLOGY — 3 LAYERS", bold=True,
          fg=WHITE, bg=NAVY, size=11, halign="center")
    row_h(ws, row, 22)

    layers = [
        ("Layer 1 — Pass/Fail Gates",
         "AUM ≥ ₹500 Cr (₹300 Cr for ELSS)  |  Fund age ≥ 3 years  |  3Y alpha > −5%"),
        ("Layer 2 — 5-Factor Quantitative Scoring",
         "Alpha (25%) · Sharpe (25%) · Consistency (25%) · Sortino (20%) · Cost (5%)"),
        ("Layer 3 — Qualitative Analyst Overlay",
         "Fund manager tenure, mandate adherence, AMC stability. Applied before final rating."),
    ]
    for i, (layer_name, layer_desc) in enumerate(layers):
        r = row + 1 + i
        ws.merge_cells(f"B{r}:D{r}")
        apply(ws.cell(r, 2), value=layer_name, bold=True, fg=WHITE,
              bg="005F73", size=10, halign="left")
        ws.merge_cells(f"E{r}:I{r}")
        apply(ws.cell(r, 5), value=layer_desc, fg=DARK_TEXT,
              bg=ZEBRA if i % 2 == 0 else WHITE, size=10, halign="left")
        row_h(ws, r, 18)

    # Column widths
    col_w(ws, 2, 3)
    for c in range(2, 10):
        if ws.column_dimensions[get_column_letter(c)].width == 0:
            col_w(ws, c, 22)
    col_w(ws, 2, 24)
    col_w(ws, 4, 10)
    col_w(ws, 5, 45)


# --- UNIVERSE SUMMARY TAB ---
def build_summary(wb, cat_results):
    ws = wb.create_sheet("Universe Summary")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"

    ws.merge_cells("A1:N1")
    apply(ws["A1"], value="UNIVERSE SUMMARY — All Equity Categories",
          bold=True, fg=WHITE, bg=NAVY, size=13, halign="center")
    row_h(ws, 1, 30)

    ws.merge_cells("A2:N2")
    apply(ws["A2"],
          value=f"Data: {DATA_DATE}  |  Regular Growth Plans  |  Primary benchmark only  |  Deduplication: first row per scheme",
          fg="555555", bg=LT_BLUE, size=9, halign="left", italic=True)
    row_h(ws, 2, 16)

    # Header
    cols = ["Category", "SEBI Mandate (Summary)", "Universe", "Post-Gates",
            "Strong Buy", "Buy", "Hold", "Underweight", "Avoid",
            "Avg Composite", "Avg 3Y Alpha", "Avg Sharpe",
            "Avg ER (%)", "Notes"]
    widths = [20, 40, 10, 10, 10, 8, 8, 12, 8, 12, 12, 10, 10, 30]
    row = 4
    for c, (h, w) in enumerate(zip(cols, widths), start=1):
        apply(ws.cell(row, c), value=h, bold=True, fg=WHITE, bg=NAVY,
              size=9, halign="center", wrap=True)
        col_w(ws, c, w)
    row_h(ws, row, 28)

    row = 5
    for cat in SCORED_CATS:
        if cat not in cat_results:
            continue
        res = cat_results[cat]
        scored = res['scored']
        n_raw  = res['n_raw']
        n_post = len(scored)

        band_counts = scored['rating'].value_counts().to_dict() if n_post > 0 else {}
        avg_comp   = scored['composite'].mean() if n_post > 0 else np.nan
        avg_alpha  = scored['alpha_3YEAR'].mean() if n_post > 0 else np.nan
        avg_sharpe = scored['RR3_Sharpe Ratio'].mean() if n_post > 0 else np.nan
        avg_er     = scored['EXR_Expense Ratio(%)'].mean() if n_post > 0 else np.nan

        notes = ""
        if cat == "Equity Linked Savings Scheme":
            notes = "3-year lock-in. Consistency weight 30% (vs 25% default)."
        elif cat == "Focused Fund":
            notes = "Max 30 stocks. Concentration risk implicit in std dev."

        bg = ZEBRA if (row % 2 == 0) else WHITE
        vals = [
            TAB_NAMES.get(cat, cat),
            SEBI_MANDATES.get(cat, ""),
            n_raw,
            n_post,
            band_counts.get("Strong Buy", 0),
            band_counts.get("Buy", 0),
            band_counts.get("Hold", 0),
            band_counts.get("Underweight", 0),
            band_counts.get("Avoid", 0),
            round(avg_comp, 1) if not np.isnan(avg_comp) else "—",
            round(avg_alpha, 2) if not np.isnan(avg_alpha) else "—",
            round(avg_sharpe, 2) if not np.isnan(avg_sharpe) else "—",
            round(avg_er, 2) if not np.isnan(avg_er) else "—",
            notes,
        ]
        for c, v in enumerate(vals, start=1):
            cell = ws.cell(row, c)
            apply(cell, value=v, fg=DARK_TEXT, bg=bg, size=9,
                  halign="center" if c > 2 else "left", wrap=(c in [2, 14]))
        row_h(ws, row, 18)
        row += 1

    # Info categories
    for cat in INFO_CATS:
        if cat not in cat_results:
            continue
        res = cat_results[cat]
        n_raw = res['n_raw']
        bg = ZEBRA if (row % 2 == 0) else WHITE
        vals = [
            TAB_NAMES.get(cat, cat),
            SEBI_MANDATES.get(cat, ""),
            n_raw, "N/A (info only)",
            "—", "—", "—", "—", "—", "—", "—", "—", "—",
            "Universe too small for percentile ranking."
        ]
        for c, v in enumerate(vals, start=1):
            cell = ws.cell(row, c)
            apply(cell, value=v, fg="888888", bg=bg, size=9,
                  halign="center" if c > 2 else "left", italic=True, wrap=(c in [2, 14]))
        row_h(ws, row, 18)
        row += 1


# --- CATEGORY TAB ---
METRIC_COLS = [
    # (header, data_col, width, number_fmt, halign)
    ("#",                   "rank",                      4,  None,    "center"),
    ("Fund Name",           "Scheme Name",               38, None,    "left"),
    ("AMC",                 "SD_AMC Full Name",           22, None,    "left"),
    ("AUM (₹ Cr)",          "AUMT_AUM",                  11, "#,##0", "right"),
    ("Age (Yrs)",           "SD_Age (From Incept Date)", 9,  "0.0",   "right"),
    ("Benchmark",           "SD_Benchmark Index",        22, None,    "left"),
    ("1Y Alpha",            "alpha_1YEAR",               9,  "+0.00;-0.00;0.00", "right"),
    ("3Y Alpha",            "alpha_3YEAR",               9,  "+0.00;-0.00;0.00", "right"),
    ("5Y Alpha",            "alpha_5YEAR",               9,  "+0.00;-0.00;0.00", "right"),
    ("Sharpe (3Y)",         "RR3_Sharpe Ratio",          10, "0.00",  "right"),
    ("Sortino (3Y)",        "RR3_Sortino",               10, "0.00",  "right"),
    ("Std Dev (3Y, %)",     "RR3_Std. Deviation",        12, "0.00",  "right"),
    ("Up Cap",              "RR3_Up Capture Ratio",      9,  "0.0",   "right"),
    ("Dn Cap",              "RR3_Down Capture Ratio",    9,  "0.0",   "right"),
    ("ER (%)",              "EXR_Expense Ratio(%)",      8,  "0.00",  "right"),
    ("ER Spread",           "er_spread",                 9,  "+0.00;-0.00;0.00", "right"),
    ("Composite",           "composite",                 10, "0.0",   "right"),
    ("Rating",              "rating",                    12, None,    "center"),
]

INFO_METRIC_COLS = [
    # Informational tabs — no rank or rating
    ("Fund Name",           "Scheme Name",               38, None,    "left"),
    ("AMC",                 "SD_AMC Full Name",           22, None,    "left"),
    ("AUM (₹ Cr)",          "AUMT_AUM",                  11, "#,##0", "right"),
    ("Age (Yrs)",           "SD_Age (From Incept Date)", 9,  "0.0",   "right"),
    ("1Y Alpha",            "alpha_1YEAR",               9,  "+0.00;-0.00;0.00", "right"),
    ("3Y Alpha",            "alpha_3YEAR",               9,  "+0.00;-0.00;0.00", "right"),
    ("5Y Alpha",            "alpha_5YEAR",               9,  "+0.00;-0.00;0.00", "right"),
    ("Sharpe (3Y)",         "RR3_Sharpe Ratio",          10, "0.00",  "right"),
    ("Sortino (3Y)",        "RR3_Sortino",               10, "0.00",  "right"),
    ("ER (%)",              "EXR_Expense Ratio(%)",      8,  "0.00",  "right"),
    ("ER Spread",           "er_spread",                 9,  "+0.00;-0.00;0.00", "right"),
]


def build_cat_tab(wb, tab_name, category, scored_df, n_raw, is_info=False):
    ws = wb.create_sheet(tab_name)
    ws.sheet_view.showGridLines = False

    col_defs = INFO_METRIC_COLS if is_info else METRIC_COLS

    # Row 1: Title
    n_cols = len(col_defs)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    display = TAB_NAMES.get(category, category)
    suffix  = " — Informational (No Scoring)" if is_info else " — Ranked Screener"
    apply(ws.cell(1, 1), value=f"{display}{suffix}",
          bold=True, fg=WHITE, bg=NAVY, size=12, halign="center")
    row_h(ws, 1, 28)

    # Row 2: Meta info
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    mandate = SEBI_MANDATES.get(category, "")
    apply(ws.cell(2, 1),
          value=f"SEBI mandate: {mandate}   |   Data: {DATA_DATE}   |   Universe: {n_raw}   |   Post-gates: {len(scored_df)}",
          fg="555555", bg=LT_BLUE, size=9, halign="left", italic=True)
    row_h(ws, 2, 15)

    # Row 3: Gate note
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=n_cols)
    if is_info:
        gate_note = "Universe too small for robust percentile ranking. Full metrics shown for reference only. No composite score or rating assigned."
    else:
        aum_t = AUM_GATE.get(category, AUM_GATE["default"])
        w = WEIGHTS.get(category, WEIGHTS["default"])
        gate_note = (
            f"Gates: AUM ≥ ₹{aum_t:,} Cr  |  Track record ≥ 3Y  |  3Y Alpha > −5%   |   "
            f"Weights: Alpha {int(w['alpha']*100)}%  Sharpe {int(w['sharpe']*100)}%  "
            f"Consistency {int(w['consistency']*100)}%  Sortino {int(w['sortino']*100)}%  "
            f"Cost {int(w['cost']*100)}%"
        )
    apply(ws.cell(3, 1), value=gate_note, fg=DARK_TEXT, bg=ZEBRA,
          size=8, halign="left", italic=True)
    row_h(ws, 3, 14)

    # Row 4: blank separator
    row_h(ws, 4, 6)

    # Row 5: Column headers
    HDR_ROW = 5
    for c, (hdr_txt, _, width, _, halign) in enumerate(col_defs, start=1):
        apply(ws.cell(HDR_ROW, c), value=hdr_txt, bold=True,
              fg=WHITE, bg=NAVY, size=9, halign="center", wrap=True)
        col_w(ws, c, width)
    row_h(ws, HDR_ROW, 24)
    ws.freeze_panes = f"A{HDR_ROW+1}"

    # Data rows
    for i, (_, row_data) in enumerate(scored_df.iterrows()):
        r = HDR_ROW + 1 + i
        bg = ZEBRA if (i % 2 == 0) else WHITE

        for c, (_, data_col, _, num_fmt, halign) in enumerate(col_defs, start=1):
            cell = ws.cell(r, c)
            raw = row_data.get(data_col) if data_col in row_data.index else None

            # Clean value
            if isinstance(raw, float) and np.isnan(raw):
                val = "—"
            elif raw is None:
                val = "—"
            else:
                val = raw

            # Rating cell: special colour
            if data_col == "rating" and val in BAND_STYLE:
                fill_hex, text_hex = BAND_STYLE[val]
                apply(cell, value=val, bold=True, fg=text_hex, bg=fill_hex,
                      size=9, halign="center", number_format=num_fmt)
            else:
                apply(cell, value=val, fg=DARK_TEXT, bg=bg,
                      size=9, halign=halign, number_format=num_fmt)

            # Colour alpha cells: green if positive, red if negative
            if data_col in ("alpha_1YEAR", "alpha_3YEAR", "alpha_5YEAR") and isinstance(val, (int, float)):
                text_c = "1a6b3c" if val >= 0 else "8b0000"
                cell.font = Font(color=text_c, size=9, name="Arial",
                                 bold=(abs(val) > 3))

        row_h(ws, r, 16)

    # Subheader band for ELSS lock-in note
    if category == "Equity Linked Savings Scheme":
        note_row = HDR_ROW + len(scored_df) + 2
        ws.merge_cells(start_row=note_row, start_column=1,
                       end_row=note_row, end_column=n_cols)
        apply(ws.cell(note_row, 1),
              value="Note: ELSS funds carry a mandatory 3-year lock-in. Consistency weight raised to 30% (vs 25% default) to reflect locked-in investor exposure.",
              fg=GOLD, bg="FFF9E6", size=9, halign="left", italic=True, border=False)


# --- CHARTS TAB (summary bar chart using openpyxl charts) ---
def build_charts_tab(wb, cat_results):
    from openpyxl.chart import BarChart, Reference
    from openpyxl.chart.label import DataLabelList

    ws = wb.create_sheet("Charts")
    ws.sheet_view.showGridLines = False

    # Data for charts stored here
    cdata = wb.create_sheet("_ChartData")
    cdata.sheet_state = "hidden"

    # --- Dataset 1: Rating distribution by category ---
    cd_row = 1
    cdata.cell(cd_row, 1).value = "Category"
    for i, b in enumerate(["Strong Buy", "Buy", "Hold", "Underweight", "Avoid"], start=2):
        cdata.cell(cd_row, i).value = b

    for cat in SCORED_CATS:
        if cat not in cat_results:
            continue
        cd_row += 1
        scored = cat_results[cat]['scored']
        cdata.cell(cd_row, 1).value = TAB_NAMES.get(cat, cat)
        counts = scored['rating'].value_counts()
        for i, b in enumerate(["Strong Buy", "Buy", "Hold", "Underweight", "Avoid"], start=2):
            cdata.cell(cd_row, i).value = int(counts.get(b, 0))

    n_cats = cd_row  # last data row for chart 1

    # --- Dataset 2: Average composite score by category ---
    cd_row += 2
    cdata.cell(cd_row, 1).value = "Category"
    cdata.cell(cd_row, 2).value = "Avg Composite Score"
    score_start_row = cd_row + 1
    for cat in SCORED_CATS:
        if cat not in cat_results:
            continue
        cd_row += 1
        scored = cat_results[cat]['scored']
        cdata.cell(cd_row, 1).value = TAB_NAMES.get(cat, cat)
        avg = scored['composite'].mean() if len(scored) > 0 else 0
        cdata.cell(cd_row, 2).value = round(avg, 1)
    score_end_row = cd_row

    # --- Dataset 3: Average 3Y Alpha by category ---
    cd_row += 2
    cdata.cell(cd_row, 1).value = "Category"
    cdata.cell(cd_row, 2).value = "Avg 3Y Alpha (%)"
    alpha_start_row = cd_row + 1
    for cat in SCORED_CATS:
        if cat not in cat_results:
            continue
        cd_row += 1
        scored = cat_results[cat]['scored']
        cdata.cell(cd_row, 1).value = TAB_NAMES.get(cat, cat)
        avg = scored['alpha_3YEAR'].mean() if len(scored) > 0 else 0
        cdata.cell(cd_row, 2).value = round(avg, 2)
    alpha_end_row = cd_row

    # --- Dataset 4: Average ER by category ---
    cd_row += 2
    cdata.cell(cd_row, 1).value = "Category"
    cdata.cell(cd_row, 2).value = "Avg ER (%)"
    er_start_row = cd_row + 1
    for cat in SCORED_CATS:
        if cat not in cat_results:
            continue
        cd_row += 1
        scored = cat_results[cat]['scored']
        cdata.cell(cd_row, 1).value = TAB_NAMES.get(cat, cat)
        avg = scored['EXR_Expense Ratio(%)'].mean() if len(scored) > 0 else 0
        cdata.cell(cd_row, 2).value = round(avg, 2)
    er_end_row = cd_row

    # Title for Charts sheet
    ws.merge_cells("A1:P1")
    apply(ws["A1"], value="CHARTS & ANALYTICS — Equity Fund Screening Model v2",
          bold=True, fg=WHITE, bg=NAVY, size=13, halign="center")
    row_h(ws, 1, 28)

    # Chart 1: Stacked bar — rating distribution
    bar = BarChart()
    bar.type     = "bar"
    bar.grouping = "stacked"
    bar.title    = "Rating Distribution by Category"
    bar.y_axis.title = "Number of Funds"
    bar.x_axis.title = "Category"
    bar.width    = 22
    bar.height   = 14

    cats_ref = Reference(cdata, min_col=1, min_row=2, max_row=n_cats)
    for col_i, band in enumerate(["Strong Buy", "Buy", "Hold", "Underweight", "Avoid"], start=2):
        data_ref = Reference(cdata, min_col=col_i, min_row=1, max_row=n_cats)
        bar.add_data(data_ref, titles_from_data=True)
    bar.set_categories(cats_ref)

    # Colour series to match rating bands
    colors = ["1a6b3c", "2e7d32", "b8860b", "c0392b", "7b0000"]
    for i, series in enumerate(bar.series):
        series.graphicalProperties.solidFill = colors[i]
    ws.add_chart(bar, "A3")

    # Chart 2: Avg composite score
    bar2 = BarChart()
    bar2.type   = "col"
    bar2.title  = "Average Composite Score by Category"
    bar2.y_axis.title = "Score (0–100)"
    bar2.width  = 22
    bar2.height = 14
    cats_ref2  = Reference(cdata, min_col=1, min_row=score_start_row, max_row=score_end_row)
    data_ref2  = Reference(cdata, min_col=2, min_row=score_start_row-1, max_row=score_end_row)
    bar2.add_data(data_ref2, titles_from_data=True)
    bar2.set_categories(cats_ref2)
    bar2.series[0].graphicalProperties.solidFill = NAVY
    ws.add_chart(bar2, "A33")

    # Chart 3: Avg 3Y alpha
    bar3 = BarChart()
    bar3.type   = "col"
    bar3.title  = "Average 3Y Alpha (%) by Category"
    bar3.y_axis.title = "Alpha (%)"
    bar3.width  = 22
    bar3.height = 14
    cats_ref3  = Reference(cdata, min_col=1, min_row=alpha_start_row, max_row=alpha_end_row)
    data_ref3  = Reference(cdata, min_col=2, min_row=alpha_start_row-1, max_row=alpha_end_row)
    bar3.add_data(data_ref3, titles_from_data=True)
    bar3.set_categories(cats_ref3)
    bar3.series[0].graphicalProperties.solidFill = "005F73"
    ws.add_chart(bar3, "M3")

    # Chart 4: Avg ER
    bar4 = BarChart()
    bar4.type   = "col"
    bar4.title  = "Average Expense Ratio (%) by Category"
    bar4.y_axis.title = "ER (%)"
    bar4.width  = 22
    bar4.height = 14
    cats_ref4  = Reference(cdata, min_col=1, min_row=er_start_row, max_row=er_end_row)
    data_ref4  = Reference(cdata, min_col=2, min_row=er_start_row-1, max_row=er_end_row)
    bar4.add_data(data_ref4, titles_from_data=True)
    bar4.set_categories(cats_ref4)
    bar4.series[0].graphicalProperties.solidFill = GOLD
    ws.add_chart(bar4, "M33")


# ══════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════

def main():
    print("Loading data...")

    # Load equity file
    eq_df = load_clean(BASE + "all Fund Screening Data.xlsx")
    eq_df = compute_derived(eq_df)

    # Load hybrid file for Focused Fund
    hy_df = load_clean(BASE + "hybrid Fund Screening Data.xlsx")
    hy_df = compute_derived(hy_df)
    focused_df = hy_df[hy_df['SD_Category'] == 'Focused Fund'].copy()

    # Combine all equity categories into one frame
    all_eq = pd.concat([eq_df, focused_df], ignore_index=True)

    all_cats = SCORED_CATS + INFO_CATS
    cat_results = {}

    total_raw    = 0
    total_scored = 0

    for cat in all_cats:
        sub = all_eq[all_eq['SD_Category'] == cat].copy()
        n_raw = len(sub)
        total_raw += n_raw

        if cat in INFO_CATS:
            # Informational — no gates, no scoring; just sort by AUM desc
            sub = sub.sort_values('AUMT_AUM', ascending=False)
            cat_results[cat] = {'scored': sub, 'n_raw': n_raw, 'n_failed': 0}
            print(f"  {cat}: {n_raw} funds (informational)")
        else:
            passed, failed = apply_gates(sub, cat)
            n_pass = len(passed)
            n_fail = len(failed)
            total_scored += n_pass

            if n_pass > 0:
                scored = score_category(passed, cat)
            else:
                scored = passed  # empty

            cat_results[cat] = {'scored': scored, 'n_raw': n_raw, 'n_failed': n_fail}
            rating_dist = scored['rating'].value_counts().to_dict() if n_pass > 0 else {}
            print(f"  {cat}: {n_raw} → {n_pass} post-gates | "
                  f"SB:{rating_dist.get('Strong Buy',0)} "
                  f"B:{rating_dist.get('Buy',0)} "
                  f"H:{rating_dist.get('Hold',0)} "
                  f"UW:{rating_dist.get('Underweight',0)} "
                  f"AV:{rating_dist.get('Avoid',0)}")

    all_stats = {'total_raw': total_raw, 'total_scored': total_scored}

    print(f"\nTotal universe: {total_raw} | Total scored: {total_scored}")
    print("Building Excel workbook...")

    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # Build tabs in order
    build_cover(wb, all_stats)
    build_summary(wb, cat_results)

    for cat in SCORED_CATS:
        tab = TAB_NAMES[cat]
        res = cat_results.get(cat, {})
        scored = res.get('scored', pd.DataFrame())
        n_raw  = res.get('n_raw', 0)
        print(f"  Building tab: {tab} ({len(scored)} funds)")
        build_cat_tab(wb, tab, cat, scored, n_raw, is_info=False)

    for cat in INFO_CATS:
        tab = TAB_NAMES[cat]
        res = cat_results.get(cat, {})
        df  = res.get('scored', pd.DataFrame())
        n_raw = res.get('n_raw', 0)
        print(f"  Building tab: {tab} ({len(df)} funds, info only)")
        build_cat_tab(wb, tab, cat, df, n_raw, is_info=True)

    build_charts_tab(wb, cat_results)

    print(f"Saving to {OUT}...")
    wb.save(OUT)
    print("Done.")


if __name__ == "__main__":
    main()
